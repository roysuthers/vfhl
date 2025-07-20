import io
import logging.config
import os
import random
import re
import requests
import smtplib
import sqlite3
import ssl
import subprocess
import sys
import textwrap
import threading
import traceback
from datetime import datetime, date, timedelta
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from inspect import currentframe, isfunction
from pathlib import Path
from textwrap import dedent
from typing import Any, Dict, List, Tuple, Union
from unidecode import unidecode

import keyring
import numpy as np
import openpyxl
import pandas as pd
import PySimpleGUI as sg
import ujson as json
from isoweek import Week
from openpyxl.utils.dataframe import dataframe_to_rows
from pandas.io.formats.style import Styler
from tabulate import tabulate

# NHL Pool classes
import get_player_data
import injuries
import player_lines
import player_stats as ps
from clsFantrax import Fantrax
from clsMoneyPuck import MoneyPuck
from clsNHL_API import NHL_API
from clsPlayer import Player
from clsPoolTeam import PoolTeam
from clsPoolTeamRoster import PoolTeamRoster
from clsSeason import Season
from clsTeam import Team
from constants import DATABASE, DATA_INPUT_FOLDER, NHL_API_URL
from utils import assign_player_ids, get_db_connection, get_db_cursor, get_player_id, load_nhl_team_abbr_and_id_dict, load_player_name_and_id_dict, split_seasonID_into_component_years

FONT = 'Consolas 12'
sg.SetOptions(
    auto_size_buttons=False,
    auto_size_text=False,
    button_color=('white', 'dark slate gray'),
    button_element_size=(8, 1),
    background_color='black',
    # dpi_awareness used to fix scaling problem after using matplotlib
    # dpi_awareness=True,
    element_background_color='black',
    element_text_color='white',
    font=FONT,
    input_elements_background_color='dark slate gray',
    input_text_color='white',
    text_color='white',
    text_element_background_color='black',
    text_justification='left',
)

# Set file path to Excel spreadsheet for current NHL season
excelFilepath = "C:\\Users\\Roy\\NHL Pool\\Excel Files\\{0}\\NHL Pool {0}.xlsx"

email_password = None

format_date = compile("lambda x: '' if x in ['', 0] else x", '<string>', 'eval')
format_int = compile("lambda x: '' if x in ['', 0] else '{:}'.format(x)", '<string>', 'eval')
format_0_decimals = compile("lambda x: '' if pd.isna(x) or x in ['', 0] else '{:0.0f}'.format(x)", '<string>', 'eval')
format_1_decimal = compile("lambda x: '' if pd.isna(x) or x in ['', 0] else '{:0.1f}'.format(x)", '<string>', 'eval')
format_2_decimals = compile("lambda x: '' if pd.isna(x) or x in ['', 0] else '{:0.2f}'.format(x)", '<string>', 'eval')
# none_to_empty = compile("lambda x:'' if x is None else x", '<string>', 'eval')
format_3_decimals_no_leading_0 = compile("lambda x: '' if pd.isna(x) or x in ['', 0] else '{:.3f}'.format(x).lstrip('0')", '<string>', 'eval')
zero_toi_to_empty = compile("lambda x:'' if x in ('00:00', None) else x", '<string>', 'eval')
format_nan_to_empty = compile("lambda x: '' if pd.isna(x) else x", '<string>', 'eval')

class HockeyPool:

###################################################################################
# New functions:
#
#   - find players traded in pre-season, and back to original manager
        # SELECT
        #     t1.Season,
        #     t1.Player,
        #     t1.[Trade #] AS Trade1,
        #     t1.Period AS Period1,
        #     t1.[From] AS From1,
        #     t1.[To] AS To_1,
        #     t2.[Trade #] AS Trade2,
        #     t2.Period AS Period2,
        #     t2.[From] AS From2,
        #     t2.[To] AS To_2
        # FROM dfHistoryTrades t1
        # JOIN dfHistoryTrades t2
        #     ON t1.Season = t2.Season
        #     AND t1.Player = t2.Player
        #     AND t1.[Trade #] < t2.[Trade #] -- Ensures no self-join
        #     AND t1.[From] = t2.[To]         -- Matches From_1 with To_2
        #     AND t1.[To] = t2.[From]         -- Matches To_1 with From_2
        # WHERE t1.Player NOT LIKE "%Draft Pick%"
        # ORDER BY t1.Season, t1.[Trade #], t1.Player;

#   - generate historical average stats for each draft pick
        # CREATE VIEW DraftedPlayerStats AS
        # SELECT
        #     dr.season_id,
        #     dr.player_id,
        #     dr.player,
        #     dr.pos,
        #     dr.[round],
        #     dr.pick,
        #     (
        #         SELECT COUNT(*)
        #         FROM dfDraftResults dr2
        #         WHERE dr2.season_id = dr.season_id AND dr2.overall <= dr.overall
        #     ) AS overall, -- Renumbering within each season_id
        #     dr.pool_team AS PickedBy,
        #     ps.games,
        #     ps.points,
        #     ps.goals,
        #     ps.assists,
        #     ps.points_pp AS ppp,
        #     ps.shots AS sog,
        #     ps.pim,
        #     ps.hits,
        #     ps.blocked AS blks,
        #     ps.takeaways AS tk,
        #     ps.games_started AS starts,
        #     ps.wins,
        #     ps.saves,
        #     round(ps.gaa, 2) AS gaa,
        #     round(ps.[save%], 3) AS [save%],
        #     ps.goals_against AS ga,
        #     ps.shots_against AS sa,
        #     (CAST(SUBSTR(ps.toi, 1, INSTR(ps.toi, ':') - 1) AS INTEGER) * 60) +
        #     CAST(SUBSTR(ps.toi, INSTR(ps.toi, ':') + 1) AS INTEGER) AS toi
        # FROM dfDraftResults dr
        # LEFT OUTER JOIN PlayerStats ps
        #     ON ps.seasonID = dr.season_id AND ps.player_id = dr.player_id AND ps.season_type = "R";
# use the following to get a list showing for each overall pick, the number of forwards, defensemen, and goalies
        # SELECT
        #     overall,
        #     SUM(CASE WHEN pos = 'F' THEN 1 ELSE 0 END) AS Forwards,
        #     SUM(CASE WHEN pos = 'D' THEN 1 ELSE 0 END) AS Defensemen,
        #     SUM(CASE WHEN pos = 'G' THEN 1 ELSE 0 END) AS Goalies
        # FROM DraftedPlayerStats
        # WHERE overall <= 132
        # GROUP BY overall
        # ORDER BY overall;

#
#   - generate top players in each scoring category, per season, using projected stats
#       - top 11 x 16 skaters
#       - top 11 x 9 forwards
#       - top 11 x 6 defense
#       - top 11 x 2 goalies
###################################################################################

    def __init__(self):

        self.id = 0
        self.name = ''
        self.season_id = 0
        self.web_host = 'Fantrax'
        self.draft_rule_id = None
        self.fee = 0
        self.league_id = ''

    def align_center(self, s):
        #  Center-align table text columns
        is_value = [True if isinstance(x, int) or isinstance(x, np.int64) else False for x in s]

        css = []
        for value in is_value:
            if value:
                css.append('text-align: center')
            else:
                css.append('text-align: left')

        return css

    def align_right(self, s):
        #  Right-align table text based on cell value
        is_value = [True if x=='Totals' else False for x in s]

        css = []
        for value in is_value:
            if value:
                css.append('text-align: right')
            else:
                css.append('text-align: left')

        return css

    def apiGetPlayerSeasonStats(self):

        try:

            NHL_API().get_player_stats(season=season)

        except Exception as e:
            sg.popup_error(f'Error in {sys._getframe().f_code.co_name}: {e}')

        return

    def apiGetSeason(self, season: Season):

        try:

            with get_db_connection() as connection:

                # target_season = f'{season.name[0:4]}{season.name[5:9]}'
                season_info = NHL_API().get_season(season=season)
                if len(season_info) == 0:
                    sg.Popup('Failed to return season information.', title='Error')
                    return

                season.start_date = re.search(r'\d{4}-\d{2}-\d{2}',season_info['regularSeasonStartDate']).group()
                season.end_date = re.search(r'\d{4}-\d{2}-\d{2}',season_info['regularSeasonEndDate']).group()
                season.number_of_games = season_info['numberOfGames']
                season.count_of_total_game_dates = season_info['totalGameDates']

                # Calculated weeks in season
                start_date = datetime.strptime(season.start_date, '%Y-%m-%d')
                end_date = datetime.strptime(season.end_date, '%Y-%m-%d')
                start_week = Week.withdate(start_date)
                end_week = Week.withdate(end_date)
                if start_date.year < end_date.year:
                    last_week_of_start_year = Week.last_week_of_year(int(season.start_date[0:4]))
                    season.weeks = last_week_of_start_year.week - start_week.week + end_week.week + 1
                else: # in same year, such as 20202021 Covid-19 schedule
                    season.weeks = end_week.week  - start_week.week+ 1

                ret = season.persist(connection)
                if ret == False:
                    sg.Popup('Failed to persist NHL season information.', title='Error')
                    return

        except Exception as e:
            msg = ''.join(traceback.format_exception(type(e), value=e, tb=e.__traceback__))
            sg.popup_error(msg)

        connection.close()

        return

    def apiGetTeams(self):

        try:

            with get_db_connection() as connection:

                # get teams from NHL api
                nhl_teams = NHL_API().get_teams(season=season)
                for row in nhl_teams.itertuples():
                    team = Team()
                    team.id = row.id
                    team.name = row.fullName
                    team.abbr = row.triCode
                    ret = team.persist(season.id, connection)
                    if ret == False:
                        sg.Popup('Failed to persist NHL team in target season.', title='Error')
                        return

        except Exception as e:
            msg = ''.join(traceback.format_exception(type(e), value=e, tb=e.__traceback__))
            sg.popup_error(msg)

        connection.close()

        return

    def archive_keeper_lists(self, season: Season, pool: 'clsHockeyPool'):

        columns = [
            f'{pool.id} as pool_id',
            'pt.name as pool_team',
            'ptr.player_id',
            'p.full_name as player_name',
            't.abbr as team_abbr',
            'p.primary_position as pos',
            'ptr.keeper'
        ]

        select_sql = ', '.join([
            ' '.join(['select', ', '.join(columns)]),
        ])

        # build table joins
        from_tables = textwrap.dedent('''\
            from PoolTeamRoster ptr
            left outer join Player p on p.id=ptr.player_id
            left outer join PoolTeam pt ON pt.id=ptr.poolteam_id
            left outer join Team t on t.id=p.current_team_id
        ''')

        where_clause = f'where pt.pool_id={pool.id} and (ptr.keeper="y" or ptr.keeper="m")'

        sql = textwrap.dedent(f'''\
            {select_sql}
            {from_tables}
            {where_clause}
        ''')

        dfKeeperLists = pd.read_sql(sql, con=get_db_connection())

        # sql = f'delete from KeeperListsArchive where season_id={season.id}'
        sql = f'delete from KeeperListsArchive where pool_id={pool.id}'
        with get_db_connection() as connection:
            connection.execute(sql)
            connection.commit()

        dfKeeperLists.to_sql('KeeperListsArchive', con=get_db_connection(), index=False, if_exists='append')

        return

    def ask_for_email_recipients(self):

        try:

            # Ask to whom the emails should be sent
            recipients = []
            # Get pool participants
            pool_teams = self.get_pool_teams_as_objects()
            for pool_team in pool_teams:
                if pool_team.email:
                    recipients.append(''.join([pool_team.name, ' (', pool_team.email, ')']))

            event, values = sg.Window('Select Users to Receive Email',layout=[[sg.Listbox(recipients,key='_LIST_',size=(max([len(str(v)) for v in recipients])+ 2,len(recipients)),select_mode='extended',bind_return_key=True),sg.OK()]]).read(close=True)
            chosen = values['_LIST_'] if event is not None else None
            if len(chosen) > 0:
                recipients.clear()
                for choice in chosen:
                    recipients.append(choice)

        except Exception as e:
            sg.popup_error(f'Error in {sys._getframe().f_code.co_name}: {e}')

        return recipients

    def askForWeekNumber(self) -> int:

        season = Season().getSeason(id=self.season_id)

        layout = [
            [
                sg.Text('Week number:', size=(15,1)),
                sg.Combo(values=[x for x in range(1, season.WEEKS_IN_NHL_SEASON + 1)], size=(5,1), default_value=season.CURRENT_WEEK, key='_WEEK_NUMBER_'),
            ],
            [
                sg.OK(),
                sg.Cancel(),
            ],
        ]

        window = sg.Window(title='NHL Pool', layout=layout)

        # Ask for stats week number
        while True:
            try:
                event, values = window.read()
                if event in (None, 'Cancel'):
                    window.close()
                    return 0
                if event == 'OK':
                    week_number = values['_WEEK_NUMBER_']
                    window.close()
                if week_number is None or week_number == '' or not (1 <= int(week_number) <= season.WEEKS_IN_NHL_SEASON):
                    raise ValueError
            except ValueError:
                sg.popup_error(f'"{week_number}"" is not valid. Must be between 1 and {season.WEEKS_IN_NHL_SEASON}', title='NHL Pool')
                continue
            else:
                break

        return int(week_number)

    def askToScrapeWebPage(self):

        layout = [
            [sg.Text(f'Do you want to scrape the data or extract from database?')],
            [sg.Text('   1. Scrape from web site')],
            [sg.Text('   2. Extract raw data from database table')],
            [
                sg.Text('Select option:', size=(15, 1)),
                sg.Input(default_text='1', justification='right', key='_OPTION_', size=(5,1), pad=(1,0), do_not_clear=False, focus=True),
            ],
            [sg.OK(), sg.Cancel()],
        ]

        dialog = sg.Window('NHL Pool', layout, font=("arial", 12), button_color=('black', 'deep sky blue'), margins=(10, 10))

        while True:
            try:
                option = ''
                event, values = dialog.Read()
                if event in (None, 'Cancel'):
                    break
                if event == 'OK':
                    option = int(values['_OPTION_'])
                    if option not in (1, 2):
                        raise ValueError('Response not valid.')
                    # website = self.web_host
                    break
                continue
            except ValueError as e:
                msg = ''.join(traceback.format_exception(type(e), value=e, tb=e.__traceback__))
                sg.popup_error(msg)
                continue
            else:
                #we're ready to exit the loop.
                break

        dialog.Close()

        return option

    # def askToUseNHL_RestService(self):

    #     layout = [
    #         [sg.Text(f'Do you want to read data from the NHL Rest Service or extract from database?')],
    #         [sg.Text('   1. Scrape from Rest Service')],
    #         [sg.Text('   2. Extract raw data from database table')],
    #         [
    #             sg.Text('Select option:', size=(15, 1)),
    #             sg.Input(justification='right', key='_OPTION_', size=(5,1), pad=(1,0), do_not_clear=False, focus=True),
    #         ],
    #         [sg.OK(), sg.Cancel()],
    #     ]

    #     dialog = sg.Window('NHL Pool', layout, font=("arial", 12), button_color=('black', 'deep sky blue'), margins=(10, 10))

    #     while True:
    #         try:
    #             option = ''
    #             event, values = dialog.Read()
    #             if event in (None, 'Cancel'):
    #                 break
    #             if event == 'OK':
    #                 option = int(values['_OPTION_'])
    #                 if option not in (1, 2):
    #                     raise ValueError('Response not valid.')
    #                 break
    #             continue
    #         except ValueError as e:
    #             print(f'{Path(__file__).stem}, Line {e.__traceback__.tb_lineno}: {e}({str(e)})')
    #             continue
    #         else:
    #             #we're ready to exit the loop.
    #             break

    #     dialog.Close()

    #     return option

    def config_pool_team(self, web_host: Tuple[str, None]=None):

        global pool_team_config

        if web_host is None:
            web_host = self.web_host

        if web_host == 'Fantrax':
            pool_team_config = {
                'columns': [
                    {'title': 'ID', 'table column': 'id', 'visible': False},
                    {'title': 'Pool ID', 'table column': 'pool_id', 'visible': False},
                    {'title': 'Name', 'table column': 'name', 'justify': 'left'},
                    {'title': 'Points', 'table column': 'points'},
                ],
                # 'order by': [{'column': 'name', 'sequence': 'asc'}]
            }
        else:
            pool_team_config = {
                'columns': [
                    {'title': 'ID', 'table column': 'id', 'visible': False},
                    {'title': 'Pool ID', 'table column': 'pool_id', 'visible': False},
                    {'title': 'Name', 'table column': 'name', 'justify': 'left'},
                    {'title': 'Email', 'table column': 'email', 'visible': False, 'justify': 'left'},
                    {'title': 'Draft', 'table column': 'draft_pos', 'visible': False, 'format': eval(format_0_decimals)},
                    {'title': 'Pts', 'runtime column': 'pool_points', 'format': eval(format_0_decimals)},
                    {'title': 'PTW', 'runtime column': 'ptw', 'format': '{:}'},
                    {'title': 'GP', 'runtime column': 'games', 'format': eval(format_0_decimals)},
                    {'title': 'PpG', 'runtime column': 'points_per_game', 'format': eval(format_2_decimals)},
                    {'title': 'Paid', 'table column': 'paid', 'format': '${:.0f}', 'sum': True},
                    {'title': 'Won', 'table column': 'won', 'format': '${:.2f}', 'sum': True},
                ],
            }

        headings = [x['title'] for x in pool_team_config['columns']]
        visible_columns = [False if ('visible' in x and x['visible'] is False) else True for x in pool_team_config['columns']]

        return (headings, visible_columns)

    def config_pool_team_roster(self, web_host: Tuple[str, None]=None):

        global pool_team_roster_config

        if web_host is None:
            web_host = self.web_host

        pool_team_roster_config = {
            'columns': [
                {'title': 'Season ID', 'table column': 'seasonID', 'visible': False},
                {'title': 'Player ID', 'table column': 'player_id', 'visible': False},
                {'title': 'Team ID', 'table column': 'team_id', 'visible': False},
                {'title': 'Name', 'table column': 'name', 'justify': 'left'},
                {'title': 'Status', 'table column': 'status'},
                {'title': 'Team', 'table column': 'team_abbr'},
                {'title': 'Pos', 'table column': 'pos'},
                {'title': 'Keeper', 'table column': 'keeper', 'format': eval(format_nan_to_empty)},
                {'title': 'GP', 'table column': 'games', 'format': eval(format_0_decimals), 'sum': True},
                {'title': 'Pts', 'table column': 'points', 'format': eval(format_0_decimals), 'sum': True},
                {'title': 'G', 'table column': 'goals', 'format': eval(format_0_decimals), 'sum': True},
                {'title': 'A', 'table column': 'assists', 'format': eval(format_0_decimals), 'sum': True},
                {'title': 'PIM', 'table column': 'pim', 'format': eval(format_0_decimals), 'sum': True},
                {'title': 'SOG', 'table column': 'shots', 'format': eval(format_0_decimals), 'sum': True},
                {'title': 'PPP', 'table column': 'points_pp', 'format': eval(format_0_decimals), 'sum': True},
                {'title': 'Hit', 'table column': 'hits', 'format': eval(format_0_decimals), 'sum': True},
                {'title': 'Blk', 'table column': 'blocked', 'format': eval(format_0_decimals), 'sum': True},
                {'title': 'Tk', 'table column': 'takeaways', 'format': eval(format_0_decimals), 'sum': True},
                {'title': 'W', 'table column': 'wins', 'format': eval(format_0_decimals), 'sum': True},
                {'title': 'GAA', 'table column': 'gaa', 'format': eval(format_2_decimals)},
                {'title': 'SV', 'table column': 'saves', 'format': eval(format_0_decimals), 'sum': True},
                {'title': 'SV%', 'table column': 'save%', 'format': eval(format_3_decimals_no_leading_0)},
                {'title': 'Injury', 'table column': 'injury_status', 'justify': 'left'},
            ],
            'order by': [
                {'column': 'points', 'sequence': 'desc'},
                {'column': 'name', 'sequence': 'asc'},
            ]
        }

        headings = [x['title'] for x in pool_team_roster_config['columns']]
        visible_columns = [False if ('visible' in x and x['visible'] is False) else True for x in pool_team_roster_config['columns']]

        return (headings, visible_columns)

    def email_nhl_team_transactions(self, batch: bool=False):

        try:

            if batch:
                logger = logging.getLogger(__name__)
                dialog = None
            else:
                # layout the progress dialog
                layout = [
                    [
                        sg.Text('Get NHL team transactions...', size=(60,2), key='-PROG-')
                    ],
                    [
                        sg.Cancel()
                    ],
                ]
                # create the dialog
                dialog = sg.Window('Get NHL team transactions', layout, finalize=True, modal=True)

            msg = 'Getting NHL team transactions from Fantrax...'
            if batch:
                logger.debug(msg)
            else:
                dialog['-PROG-'].update(msg)
                event, values = dialog.read(timeout=2)
                if event == 'Cancel' or event == sg.WIN_CLOSED:
                    return

            fantrax = Fantrax(pool_id=self.id, league_id=self.league_id, season_id=self.season_id)
            dfNHLTeamTransactions, dfNHLExcludeTeamTrans = fantrax.scrapeNHLTeamTransactions(dialog=dialog)
            del fantrax

            if len(dfNHLTeamTransactions.index) == 0 and len(dfNHLExcludeTeamTrans.index) == 0:
                msg = 'No NHL team transactions. Returning...'
                if batch:
                    logger.debug(msg)
                else:
                    sg.popup_notify(msg, title=sys._getframe().f_code.co_name)
                return

            ####################################################################################################
            # write dfNHLExcludeTeamTrans to database
            dfNHLExcludeTeamTrans.to_sql('dfNHLExcludeTeamTrans', con=get_db_connection(), index=False, if_exists='append')

            ####################################################################################################
            # Email transactions
            # bypass transactions that were already recorded
            dfCurrentTransactions = pd.read_sql('select * from dfNHLTeamTransactions', con=get_db_connection())
            # Perform a left-join, eliminating duplicates in df2 so that each row of df1 joins with exactly 1 row of df2.
            # Use the parameter indicator to return an extra column indicating which table the row was from.
            df_all = dfNHLTeamTransactions.merge(dfCurrentTransactions, on=['player_name', 'pos', 'team_abbr', 'comment'], how='left', indicator=True)
            data = df_all[df_all['_merge'] == 'left_only']

            # write to database
            msg = f'{len(dfNHLTeamTransactions.index)} NHL Team transactions collected. Writing to database...'
            if batch:
                logger.debug(msg)
            else:
                dialog['-PROG-'].update(msg)
                event, values = dialog.read(timeout=2)
                if event == 'Cancel' or event == sg.WIN_CLOSED:
                    return

            # save lastest transactions to database
            dfNHLTeamTransactions.to_sql('dfNHLTeamTransactions', con=get_db_connection(), index=False, if_exists='replace')

            if len(data.index) == 0:
                msg = 'No new NHL team transactions. Returning...'
                if batch:
                    logger.debug(msg)
                else:
                    sg.popup_notify(msg, title=sys._getframe().f_code.co_name)
                return

            data['player_name'] = data.apply(lambda x: self.make_clickable(column='player_name', value=x['player_name']), axis='columns')
            data.rename(columns={'player_name': 'name', 'team_abbr': 'team'}, inplace=True)
            data.drop(columns='_merge', inplace=True)

            msg = 'Preparing email html...'
            if batch:
                logger.debug(msg)
            else:
                dialog['-PROG-'].update(msg)
                event, values = dialog.read(timeout=2)
                if event == 'Cancel' or event == sg.WIN_CLOSED:
                    return

            caption = f'New NHL Transactions'

            styler = Styler(data)
            styler = styler.set_caption(caption)
            styler = styler.set_table_styles(self.setCSS_TableStyles())
            styler = styler.set_table_attributes('style="border-collapse:collapse"')

            htmlTable = styler.hide(axis="index").to_html()

            recipients = ['rsuthers@cogeco.ca']
            # recipients = ['rsuthers@cogeco.ca', 'jfjasonfowler@gmail.com']

            msg = 'Formatting & sending email...'
            if batch:
                logger.debug(msg)
                dialog = None
            else:
                dialog['-PROG-'].update(msg)
                event, values = dialog.read(timeout=2)
                if event == 'Cancel' or event == sg.WIN_CLOSED:
                    return

            email_sent = self.formatAndSendEmail(html_tables=[htmlTable], message='', recipients=recipients, subject=caption, show_sent_msg=False, batch=batch, dialog=dialog)

            if email_sent is True:
                msg = 'Email sent...'
                if batch:
                    logger.debug(msg)
                else:
                    dialog['-PROG-'].update(msg)
                    event, values = dialog.read(timeout=2)
                    if event == 'Cancel' or event == sg.WIN_CLOSED:
                        return

                sg.popup_notify('Email sent to:\n\t{0}'.expandtabs(4).format("\n\t".expandtabs(4).join(recipients)), title=sys._getframe().f_code.co_name)

                msg = 'Saving lastest transactions to datahase...'
                if batch:
                    logger.debug(msg)

        except Exception as e:
            if batch:
                logger.error(repr(e))
                raise
            else:
                sg.popup_error(f'Error in {sys._getframe().f_code.co_name}: {e}')

        finally:
            msg = 'Email NHL team transactions completed...'
            if batch:
                logger.debug(msg)
            else:
                dialog.close()
                sg.popup_notify(msg, title=sys._getframe().f_code.co_name)

        return

    def email_starting_goalie_projections(self, pool_id: int, batch: bool=False):

        try:

            if batch:
                logger = logging.getLogger(__name__)
                dialog = None
            else:
                # layout the progress dialog
                layout = [
                    [
                        sg.Text('Get Starting Goalie Projections...', size=(60,2), key='-PROG-')
                    ],
                    [
                        sg.Cancel()
                    ],
                ]
                # create the dialog
                dialog = sg.Window('Starting Goalie Projections', layout, finalize=True, modal=True)

            msg = 'Getting Starting Goalie Projections from 5v5Hockey.com...'
            if batch:
                logger.debug(msg)
            else:
                dialog['-PROG-'].update(msg)
                event, values = dialog.read(timeout=2)
                if event == 'Cancel' or event == sg.WIN_CLOSED:
                    return

            df5v5HockeyGoalieTool = player_lines._5v5_hockey_goalies(pool_id=pool_id, dialog=dialog, batch=batch)

            if len(df5v5HockeyGoalieTool.index) == 0:
                msg = 'No Starting Goalie Projections.' # used in `finally` block
                return

            dfCurrentProjections = pd.read_sql('select * from df5v5HockeyGoalieTool', con=get_db_connection())

            # Create temporary dataframes for both df5v5HockeyGoalieTool & dfCurrentProjections
            # that include only rows in which the 'Manager' column starts with 'Banshee' or 'FF', or is blank.
            temp_df5v5HockeyGoalieTool = df5v5HockeyGoalieTool[df5v5HockeyGoalieTool['Manager'].str.startswith(('Banshee', 'FF')) | df5v5HockeyGoalieTool['Manager'].eq('')]
            temp_dfCurrentProjections = dfCurrentProjections[dfCurrentProjections['Manager'].str.startswith(('Banshee', 'FF')) | dfCurrentProjections['Manager'].eq('')]

            # Sort both dataframes by 'Goalie' and reset their indexes
            temp_df5v5HockeyGoalieTool.sort_values(by='Goalie', inplace=True)
            temp_df5v5HockeyGoalieTool.reset_index(drop=True, inplace=True)

            temp_dfCurrentProjections.sort_values(by='Goalie', inplace=True)
            temp_dfCurrentProjections.reset_index(drop=True, inplace=True)

            # If the dataframes are identical, return
            if temp_df5v5HockeyGoalieTool[['Goalie', 'Manager', 'Status']].equals(temp_dfCurrentProjections[['Goalie', 'Manager', 'Status']]):
            # if df5v5HockeyGoalieTool[['Goalie', 'Status']].equals(dfCurrentProjections[['Goalie', 'Status']]):
                msg = 'No new Starting Goalie Projections. Returning...'
                if batch:
                    logger.warn(msg)
                else:
                    dialog['-PROG-'].update(msg)
                    event, values = dialog.read(timeout=2)
                return

            # Check if any rows in dfCurrentProjections have the same 'Score' that is not 0
            duplicate_scores = temp_df5v5HockeyGoalieTool[temp_df5v5HockeyGoalieTool['Score'] != 0]['Score'].duplicated(keep=False)

            if duplicate_scores.any():
                duplicate_rows = temp_df5v5HockeyGoalieTool[duplicate_scores]
                msg = f"Duplicate scores found:\n{duplicate_rows[['Goalie', 'Score']]}. Returning..."
                if batch:
                    logger.warn(msg)
                else:
                    sg.popup_notify(msg, title='Duplicate Scores Found')
                return

            ##############################################################################
            # get todays games
            dfTodaysGames = player_lines._5v5_todays_games(dialog=dialog, batch=batch)
            if len(dfTodaysGames.index) == 0:
                msg = 'No games collected for today. Returning...'
                if dialog:
                    dialog['-PROG-'].update(msg)
                    event, values = dialog.read(timeout=2)
                else:
                    logger.warn(msg)
                return

            msg = f'{len(df5v5HockeyGoalieTool.index)} New NHL Starting Goalie Projections collected. Writing to database...'
            if batch:
                logger.debug(msg)
            else:
                dialog['-PROG-'].update(msg)
                event, values = dialog.read(timeout=2)
                if event == 'Cancel' or event == sg.WIN_CLOSED:
                    return

            # flag Team as true or false, when it is the Home Team
            # df['Home Team'] = df['Team'].isin(dfTodaysGames['Home Team'])
            df5v5HockeyGoalieTool['Team'] = df5v5HockeyGoalieTool['Team'].apply(lambda team: f"@{team}" if team in dfTodaysGames['Home Team'].values else team)
            df5v5HockeyGoalieTool['Opponent'] = df5v5HockeyGoalieTool['Opponent'].apply(lambda team: f"@{team}" if team in dfTodaysGames['Home Team'].values else team)

            # save lastest transactions to database
            df5v5HockeyGoalieTool.to_sql('df5v5HockeyGoalieTool', con=get_db_connection(), index=False, if_exists='replace')

            ##############################################################################
            # reduce df5v5HockeyGoalieTool to games of interest to me and Jason

            # Step 1: Create a boolean mask for rows where Manager is empty/missing
            # or starts with "FF" or "Banshee". (Note: using na=False in str.startswith to handle NaN)
            manager_mask = (
                df5v5HockeyGoalieTool["Manager"].isna() |
                (df5v5HockeyGoalieTool["Manager"] == "") |
                df5v5HockeyGoalieTool["Manager"].str.startswith("FF", na=False) |
                df5v5HockeyGoalieTool["Manager"].str.startswith("Banshee", na=False)
            )

            # Step 2: Get the unique team names from rows satisfying the manager condition.
            teams_to_include = df5v5HockeyGoalieTool.loc[manager_mask, "Team"].unique()

            # Step 3: Filter the entire dataframe to include rows where either 'Team' or 'Opponent' is in our list.
            reduced_df = df5v5HockeyGoalieTool[
                df5v5HockeyGoalieTool["Team"].isin(teams_to_include) |
                df5v5HockeyGoalieTool["Opponent"].isin(teams_to_include)
            ]

            game_times = []
            response = requests.get(f'{NHL_API_URL}/schedule/{datetime.now().strftime("%Y-%m-%d")}')
            if response.status_code == 200:
                games = response.json()['gameWeek'][0]['games']
                game_times_utc = [{'homeTeam': game['homeTeam']['abbrev'], 'awayTeam': game['awayTeam']['abbrev'], 'startTimeUTC': game['startTimeUTC'], 'UTCOffset': game['easternUTCOffset']} for game in games]
                for game in game_times_utc:
                    homeTeam = game['homeTeam']
                    awayTeam = game['awayTeam']
                    startTimeUTC = game['startTimeUTC']
                    UTCOffset = game['UTCOffset']
                    # Parse the UTC time
                    utc_time = datetime.strptime(startTimeUTC, '%Y-%m-%dT%H:%M:%SZ')
                    # Calculate the offset as a timedelta
                    hours_offset, minutes_offset = map(int, UTCOffset.split(':'))
                    offset = timedelta(hours=hours_offset, minutes=minutes_offset)
                    # Apply the offset to get local time
                    local_time = utc_time + offset
                    # Format the local time as 'h:mm AM/PM'
                    formatted_time = local_time.strftime('%I:%M %p').lstrip('0')  # Remove leading zero from hour
                    game_times.append({'Team': homeTeam, 'Start Time': formatted_time, 'DateTime': local_time})
                    game_times.append({'Team': awayTeam, 'Start Time': formatted_time, 'DateTime': local_time})

           # Add 'Start Time' to the reduced_df dataframe
            def get_start_time(team_name):
                # Strip '@' prefix if present
                team_name = team_name.lstrip('@')
                # Find the start time for the team
                return next((item['DateTime'] for item in game_times if item['Team'] == team_name), None)

            # Apply the function to the 'Team' column to create a new 'Start Time' column
            reduced_df['DateTime'] = reduced_df['Team'].apply(get_start_time)

            reduced_df.sort_values(by=['DateTime', 'Game'], inplace=True)

            # Filter out rows where DateTime is less than the current time
            current_time = datetime.now()
            reduced_df = reduced_df[reduced_df['DateTime'] > current_time]

            if len(reduced_df.index) == 0:
                msg = 'No Starting Goalie Projections.' # used in `finally` block
                return

            reduced_df.drop(columns=['DateTime'], inplace=True)

            reduced_df.reset_index(drop=True, inplace=True)

            ##############################################################################
            msg = 'Preparing email html...'
            if batch:
                logger.debug(msg)
            else:
                dialog['-PROG-'].update(msg)
                event, values = dialog.read(timeout=2)
                if event == 'Cancel' or event == sg.WIN_CLOSED:
                    return

            recipients = ['rsuthers@cogeco.ca']
            # recipients = ['rsuthers@cogeco.ca', 'jfjasonfowler@gmail.com']

            subject = f'Starting Goalie Projections'

            # Iterating through unique 'Game' values
            unique_games = reduced_df['Game'].unique()  # Get unique 'Game' values

            html_tables = []
            new_column_order = ['Goalie', 'Team', 'Score', 'Rank', 'Manager', 'Status', 'Win%', 'xGA', 'xSA', 'xSV', 'xSV%', 'QS%', 'SV%', 'GAA']
            columns_to_center = ['Team', 'Score', 'Rank', 'Manager', 'Status', 'Win%', 'xGA', 'xSA', 'xSV', 'xSV%', 'QS%', 'SV%', 'GAA']
            for game in unique_games:

                temp_df = reduced_df[reduced_df['Game'] == game]  # Filter the dataframe for each unique 'Game'

                # Create the caption dynamically from the 'Team' column
                unique_teams = temp_df['Team'].unique()

                home_team = f"{unique_teams[1].replace('@', '')}" if unique_teams[1].startswith('@') else f"{unique_teams[0].replace('@', '')}"

                start_time = next((item['Start Time'] for item in game_times if item['Team'] == home_team), None)

                caption = f"{unique_teams[0]} @ {unique_teams[1].replace('@', '')}" if unique_teams[1].startswith('@') else f"{unique_teams[1]} @ {unique_teams[0].replace('@', '')}"
                caption = f'{caption} ({start_time})'

                # temp_df.drop(columns=['Game', 'Team', 'Opponent'], inplace=True)  # Drop the 'Game' column
                temp_df.drop(columns=['Game', 'Opponent'], inplace=True)  # Drop the 'Game' column

                temp_df = temp_df[new_column_order]

                styler = Styler(temp_df)
                styler = styler.set_caption(caption)
                styler = styler.set_table_styles(self.setCSS_TableStyles())
                styler = styler.set_table_attributes('style="border-collapse:collapse"')

                # Apply centering to the columns
                styler = styler.set_properties(**{'text-align': 'center'}, subset=columns_to_center)

                htmlTable = styler.hide(axis="index").to_html()

                html_tables.append(htmlTable)

            email_sent = self.formatAndSendEmail(html_tables=html_tables, message='', recipients=recipients, subject=subject, show_sent_msg=False, batch=batch, dialog=dialog)

            if email_sent is True:
                msg = 'Email sent...'
                if batch:
                    logger.debug(msg)
                else:
                    dialog['-PROG-'].update(msg)
                    event, values = dialog.read(timeout=2)
                    if event == 'Cancel' or event == sg.WIN_CLOSED:
                        return

                sg.popup_notify('Email sent to:\n\t{0}'.expandtabs(4).format("\n\t".expandtabs(4).join(recipients)), title=sys._getframe().f_code.co_name)

        except Exception as e:
            if batch:
                logger.error(repr(e))
            else:
                msg = ''.join(traceback.format_exception(type(e), value=e, tb=e.__traceback__))
                dialog['-PROG-'].update(msg)
                event, values = dialog.read(timeout=2)

        finally:
            if msg != 'No Starting Goalie Projections.':
                msg = 'Email Starting Goalie Projections completed...'
            if batch:
                if msg == 'No Starting Goalie Projections.':
                    logger.warn(msg)
                else:
                    logger.debug(msg)
            else:
                dialog.close()
                sg.popup_notify(msg, title=sys._getframe().f_code.co_name)

        return

    def export_to_excel(self, mclb: sg.Table):

        filename = excelFilepath.format(season.id)

        df = pd.DataFrame(data=mclb.Values, columns=mclb.ColumnHeadings)

        df.to_excel(filename, sheet_name='New Master List', index=False)

        os.system(' '.join(['start excel.exe', ''.join(['"', filename, '"'])]))

        return

    def fetch(self, **kwargs):

        sql = f'select * from HockeyPool'
        values = []
        if kwargs and 'Criteria' in kwargs:
            count = 0
            for criterion in kwargs['Criteria']:
                (property, opcode, value) = criterion
                if count > 0:
                    sql = f'{sql} and {property} {opcode} ?'
                else:
                    sql = f'{sql} where {property} {opcode} ?'
                if type(value) == str:
                    sql = f'{sql} COLLATE NOCASE'
                values.append(value)
                count += 1

        try:
            cursor = get_db_cursor()
            cursor.execute(sql, tuple(values))
            row = cursor.fetchone()

            if row:
                self.id = row['id']
                self.name = row['name']
                self.season_id = row['season_id']
                self.web_host = row['web_host']
                # self.DraftSeasonID = row['DraftSeasonID']
                # self.NumberOfPoolTeams = row['NumberOfPoolTeams']
                # self.PlayersPerPoolTeam = row['PlayersPerPoolTeam']
                # self.Fee = row['Fee']
                self.draft_rule_id = row['draft_rule_id']
                self.fee = row['fee']
                self.league_id = row['league_id']

        except sqlite3.Error as e:
            msg = ''.join(traceback.format_exception(type(e), value=e, tb=e.__traceback__))
            sg.popup_error(msg)
        except Exception as e:
            msg = ''.join(traceback.format_exception(type(e), value=e, tb=e.__traceback__))
            sg.popup_error(msg)
        finally:
            cursor.close()

        return self

    def fetch_many(self, **kwargs):

        sql = f'select * from HockeyPool'
        values = []
        if kwargs and 'Criteria' in kwargs:
            count = 0
            for criterion in kwargs['Criteria']:
                (property, opcode, value) = criterion
                if count > 0:
                    sql = f'{sql} and {property} {opcode} ?'
                else:
                    sql = f'{sql} where {property} {opcode} ?'
                if type(value) == str:
                    sql = f'{sql} COLLATE NOCASE'
                values.append(value)
                count += 1
        sql = f'{sql} order by name'

        try:
            cursor = get_db_cursor()
            cursor.execute(sql, tuple(values))
            rows = cursor.fetchall()

        except sqlite3.Error as e:
            msg = ''.join(traceback.format_exception(type(e), value=e, tb=e.__traceback__))
            sg.popup_error(msg)
        except Exception as e:
            msg = ''.join(traceback.format_exception(type(e), value=e, tb=e.__traceback__))
            sg.popup_error(msg)
        finally:
            cursor.close()

        return rows

    def formatAndSendEmail(self,
            html_tables: List[str],
            recipients: List[str],
            subject: str,
            headers: Union[str, Dict[str, str], List[str]] = 'firstrow',
            message: str='',
            title: str='Format and Send Email',
            bcc: List[str]=[],
            footer: str='',
            show_sent_msg: bool=True,
            batch: bool=False,
            dialog = None
        ):

        global email_password

        if batch:
            logger = logging.getLogger(__name__)

        # sender = 'roy.suthers@gmail.com'
        sender = 'rsuthers@cogeco.ca'
        email_password = keyring.get_password("Hockey Pool", sender)

        # Get email addresses for recipients
        to = []
        for email_address in recipients:
            match = re.match(r'(?:.+ ){0,1}\({0,1}(.+@.+\.[^)]+)\){0,1}', email_address)
            if match:
                email_address = match[1]
            else:
                sg.popup_error(f'Cannot locate email address in "{email_address}"', title='Format and Send Email')
                continue
            if email_address != '' and '@' in email_address:
                to.append(email_address)

        bcc = [sender] if (len(bcc) == 0 and sender not in to) else bcc

        # Add line breaks to prevent html from incorrectly breaking after 988 characters
        for i, _ in enumerate(html_tables):
            html_tables[i] = html_tables[i].replace('#T_', '\n#T_')\
                                           .replace('<style ', '\n<style ')\
                                           .replace('<table ', '\n<table ')\
                                           .replace('<thead>', '\n<thead>')\
                                           .replace('<tr>', '\n<tr>')\
                                           .replace('<th ', '\n<th ')\
                                           .replace('<tbody>', '\n<tbody>')\
                                           .replace('<td ', '\n<td ')

        if len(html_tables) == 0:
            text = message
            html = f'{message}\n<br /><pre>{footer}</pre>'

        elif len(html_tables) >= 1:  # Handles one or more tables dynamically
            tables = []
            for i, table_data in enumerate(html_tables):
                table = tabulate(table_data, headers=headers, tablefmt="grid", numalign="center", showindex=True)
                tables.append(table)

            # Join all tables with two newlines in between
            combined_tables = "\n\n".join(tables)
            text = f"{message}\n{combined_tables}"

            # Process message for HTML
            if message == '\n':
                message = ''
            else:
                message = message.replace('\n', '<br />')

            # Prepare the HTML version
            html_tables_str = "\n".join([str(table_data) for table_data in html_tables])
            html = f'{message}\n{html_tables_str}\n<br /><pre>{footer}</pre>'

        msg = MIMEMultipart("alternative", None, [MIMEText(text), MIMEText(html,'html')])

        # msg = MIMEText(body)
        msg['Subject'] = subject
        msg['From'] = sender
        msg['To'] = ", ".join(to)
        if bcc:
            to = to + bcc # Add bcc after the statment above, to keep it hidden

        ##################################################
        # Send email
        ##################################################

        try:

            email_sent = False

            message = 'Sending email...'
            if batch:
                logger.debug(message)
                dialog = None
            else:
                dialog['-PROG-'].update(message)
                event, values = dialog.read(timeout=2)
                if event == 'Cancel' or event == sg.WIN_CLOSED:
                    return email_sent

            # smtp_host = 'smtp.gmail.com'
            smtp_host = 'smtp.cogeco.ca'

            port = 465 # For SMTP_SSL()
            # port = 578 # For .starttls()

            # Must use gmail, it seems. I tried cogeco, but it seems to not support html messages.
            # server = smtplib.SMTP('smtp.gmail.com', 587)
            # server = smtplib.SMTP('smtp.cogeco.ca', 587)
            # server.ehlo()
            # server.starttls()
            # server.login(sender, email_password)
            # server.sendmail(sender, to, msg.as_string())

            # Create a secure SSL context
            context = ssl.create_default_context()
            # ...send email(s)
            with smtplib.SMTP_SSL(smtp_host, port, context=context) as server:

                while True:
                    try:
                        # No need to ask for password if passed in
                        if email_password in ('', None):
                            if batch is True:
                               logger.debug('Email password is required...')
                               return email_sent
                            else:
                                email_password = sg.PopupGetText("Enter password:", title='Send Email for Pool Standings', password_char='*', size=(20,1), modal=True)
                                if email_password is None:
                                    sg.popup_error(f'Cancel button used when expecting password. Exiting...', title='Format and Send Email')
                                    return email_sent
                                if email_password == '':
                                    sg.popup_error(f'No password entered. Try again...', title='Format and Send Email')
                                    continue

                        try:
                            server.login(sender, email_password)
                            break
                        except smtplib.SMTPAuthenticationError as e:
                            message = f'{Path(__file__).stem}, Line {e.__traceback__.tb_lineno}: SMTPAuthenticationError: {str(e)}'
                            if batch:
                                logger.debug(message)
                                dialog = None
                            else:
                                sg.PopupScrolled(message, title='Format and Send Email', modal=True)
                                email_password = None
                            # continue
                            return email_sent

                    except UnboundLocalError as e:
                        # if 'email_password' not in str(e):
                        message = f'{Path(__file__).stem}, Line {e.__traceback__.tb_lineno}: UnboundLocalError({str(e)})'
                        if batch:
                            logger.debug(message)
                            dialog = None
                        else:
                            sg.PopupScrolled(message, title='Format and Send Email', modal=True)
                            email_password = None
                        # continue
                        return email_sent
                    except Exception as e:
                        message = f'{Path(__file__).stem}, Line {e.__traceback__.tb_lineno}: {e}({str(e)})'
                        if batch:
                            logger.debug(message)
                            dialog = None
                        else:
                            sg.PopupScrolled(message, title='Format and Send Email', modal=True)
                            email_password = None
                        # continue
                        return email_sent

                message = 'Sending email...'
                if batch:
                    logger.debug(message)
                else:
                    dialog['-PROG-'].update(message)
                    event, values = dialog.read(timeout=2)
                    if event == 'Cancel' or event == sg.WIN_CLOSED:
                        return email_sent

                server.sendmail(sender, to, msg.as_string())
                email_sent = True

            if show_sent_msg is True:
                sg.popup_notify('Email sent to:\n\t{0}'.expandtabs(4).format("\n\t".expandtabs(4).join(recipients)), title=sys._getframe().f_code.co_name)

            server.close()

        except Exception as e:
            message = f'{Path(__file__).stem}, Line {e.__traceback__.tb_lineno}: {e}({str(e)})'
            if batch:
                logger.debug(msg)
                dialog = None
            else:
                sg.PopupScrolled(msg, title='Format and Send Email', modal=True)
                email_password = None

        return email_sent

    def getMCLBData(self, config: Dict, df: pd.DataFrame, sort_override: bool=False, return_df: bool=False) -> Union[List, pd.DataFrame]:

        table_columns = [x['table column'] if 'table column' in x else x['runtime column'] for x in config['columns'] if 'table column' in x or 'runtime column' in x]

        total_label_column = [x['table column'] if 'table column' in x else x['runtime column'] for x in config['columns'] if ('visible' not in x or ('visible' in x and x['visible'] is True)) and ('sum' not in x or ('sum' in x and x['sum'] is False))][0]

        sum_columns = [x['table column'] if 'table column' in x else x['runtime column'] for x in config['columns'] if 'sum' in x and x['sum'] is True]

        df_temp = df.copy(deep=True)

        # sorting
        if sort_override is False:
            if 'order by' in config:
                by = []
                order = []
                for s in config['order by']:
                    col = s['column']
                    seq = s['sequence']
                    if col in df_temp.columns.values:
                        # prior to sorting, set blank numeric values to 0
                        df_temp[col] = df_temp[col].apply(lambda x: 0 if x=='' else x)
                    else:
                        # can't sort if column not in dataframe
                        continue
                    by.append(col)
                    order.append(True if seq=='asc' else False)
                df_temp.sort_values(by=by, ascending=order, inplace=True)

        # add "Totals" row to end
        if len(sum_columns) > 0:
            s_temp = pd.Series(dtype='float64')
            for column in df_temp.columns.values:
                if column == total_label_column:
                    s_temp[column] = 'Totals'
                elif column in sum_columns:
                    df_temp[column] = pd.to_numeric(df_temp[column], errors='coerce').astype('Int64')
                    s_temp[column] = df_temp[column].sum()
                else:
                    s_temp[column] = ''
            df_totals = pd.concat([pd.DataFrame(), s_temp], axis=1).T
            df_temp = pd.concat([df_temp, df_totals], ignore_index=True)

        df_temp.fillna(0, inplace=True)

        df_temp = df_temp.reindex(columns=table_columns)

        # data = df_temp.astype(str).values.tolist()
        data = df_temp.values.tolist()
        if len(data) == 0:
            data = ['' for col in table_columns]

        if return_df is True:
            return df_temp

        return data

    def getMoneyPuckData(self, season: Season, batch: bool=False):

        try:

            if batch:
                logger = logging.getLogger(__name__)
                dialog = None
            else:
                # layout the progress dialog
                layout = [
                    [
                        sg.Text('Download MoneyPuck data...', size=(60,2), key='-PROG-')
                    ],
                    [
                        sg.Cancel()
                    ],
                ]
                # create the dialog
                dialog = sg.Window('Download MoneyPuck data', layout, finalize=True, modal=True)

            msg = 'Downloading Skater data from MoneyPuck...'
            if batch:
                logger.debug(msg)
            else:
                dialog['-PROG-'].update(msg)
                event, values = dialog.read(timeout=2)
                if event == 'Cancel' or event == sg.WIN_CLOSED:
                    return

            moneyPuck = MoneyPuck(season=season)
            moneyPuck.downloadData(season=season, dialog=dialog, batch=batch)
            del moneyPuck

        except Exception as e:
            if batch:
                logger.error(repr(e))
                raise
            else:
                sg.popup_error(f'Error in {sys._getframe().f_code.co_name}: {e}')

        finally:
            msg = 'MoneyPuck download completed...'
            if batch:
                logger.debug(msg)
            else:
                dialog.close()
                sg.popup_notify(msg, title=sys._getframe().f_code.co_name)

        return

    def get_pool_standings(self, batch: bool=False):

        try:

            season = Season().getSeason(id=self.season_id)

            # Input week number
            week_number = str(self.askForWeekNumber())

            # if there are rows for this week, delete them
            sql = f'delete from PoolTeamPointsByWeek where pool_id={self.id} and week={week_number}'
            with get_db_connection() as connection:
                connection.execute(sql)

                # get points per week
                points_by_week: pd.DataFrame = pd.read_sql(f'select * from PoolTeamPointsByWeek where pool_id={self.id}', con=connection)

                # get max index value
                idx1 = max(points_by_week.index)

                # global df_roster_player_stats has the data I need, I think
                # NOTE: need to filter out the "Totals" row
                points_this_week: pd.DataFrame = df_roster_player_stats.copy(deep=True).groupby(['poolteam_id']).aggregate({'pool_points': 'sum', 'ptw': 'sum'}).query('index!=""')
                points_this_week.sort_index(inplace=True)
                points_this_week.reset_index(inplace=True)
                points_this_week['poolteam_id'] = points_this_week['poolteam_id'].astype(int)
                points_this_week['ptw'] = points_this_week['ptw'].astype(int)

                for idx2 in points_this_week.index:
                    idx1 += 1
                    points_by_week.loc[idx1, 'pool_id'] = self.id
                    points_by_week.loc[idx1, 'poolteam_id'] = points_this_week.loc[idx2, 'poolteam_id']
                    points_by_week.loc[idx1, 'week'] = week_number
                    points_by_week.loc[idx1, 'points'] = points_this_week.loc[idx2, 'ptw']

                # set column data types
                points_by_week['pool_id'] = points_by_week['pool_id'].astype(int)
                points_by_week['poolteam_id'] = points_by_week['poolteam_id'].astype(int)
                points_by_week['week'] = points_by_week['week'].astype(int)
                points_by_week['points'] = points_by_week['points'].astype(int)

                # write points_by_week to database table
                points_by_week.to_sql('PoolTeamPointsByWeek', con=connection, if_exists='replace', index=False)

                # read pivot table
                points_by_week_pivot: pd.DataFrame = pd.read_sql(sql=f'select * from PoolTeamPointsByWeekPivot where pool_id={self.id}', con=connection)

                # Find poolies with maximum points for week
                poolie_winners_this_week: List[int] = list(points_this_week.loc[points_this_week['ptw']==points_this_week['ptw'].max()]['poolteam_id'].values)

                amount_won_per_winner = int(((self.fee * len(points_this_week.index)) / season.weeks) / len(poolie_winners_this_week))

                # insert columns at desired locations
                if f'{week_number}' not in list(points_by_week_pivot.columns):
                    points_by_week_pivot.insert(points_by_week_pivot.columns.get_loc(f'${int(week_number) - 1}') + 1, f'{week_number}', 0)
                    points_by_week_pivot.insert(points_by_week_pivot.columns.get_loc(f'${int(week_number) - 1}') + 2, f'${week_number}', 0)

                for idx in points_by_week_pivot.index:
                    poolteam_id = points_by_week_pivot.loc[idx, 'poolteam_id']
                    kwargs = {'Criteria': [['id', '==', poolteam_id], ['pool_id', '==', self.id]]}
                    poolTeam = PoolTeam().fetch(**kwargs)
                    points_by_week_pivot.loc[idx, f'{week_number}'] = points_by_week.query('poolteam_id==@poolteam_id and week==@week_number')['points'].values[0]
                    if poolteam_id in poolie_winners_this_week:
                        points_by_week_pivot.loc[idx, f'${week_number}'] = amount_won_per_winner
                    else:
                        points_by_week_pivot.loc[idx, f'${week_number}'] = 0
                    # update pool teams with amount won
                    poolTeam.won = points_by_week_pivot.loc[idx, points_by_week_pivot.columns.str.startswith('$')].sum(axis=0)
                    poolTeam.persist(connection=connection)

                # add total pool points; not a sum of the weeks because point adjustments during the season
                for idx in points_this_week.index:
                    poolteam_id: int = points_this_week.loc[idx, 'poolteam_id']
                    idx2 = points_by_week_pivot.query('pool_id==@self.id and poolteam_id==@poolteam_id').index.values[0]
                    points_by_week_pivot.loc[idx2, 'Total Points'] = points_this_week.loc[idx, 'pool_points']

                # add total won
                points_by_week_pivot['Total Won'] = points_by_week_pivot.loc[:, points_by_week_pivot.columns.str.startswith('$')].sum(axis=1)

                points_by_week_pivot[week_number] = points_by_week_pivot[week_number].astype(int)
                points_by_week_pivot.fillna(0, inplace=True)
                points_by_week_pivot.sort_values(['Total Points', 'Name'], ascending=[False, True], inplace=True)
                points_by_week_pivot.reset_index(drop=True, inplace=True)

                # set column data types
                points_by_week_pivot['Total Points'] = points_by_week_pivot['Total Points'].astype(int)

                # for final week, add unallocated entry fees to pool winner
                final_week = [x for x in range(1, season.WEEKS_IN_NHL_SEASON + 1)][-1]
                if int(week_number) == final_week:
                    # Find poolies with maximum points for season
                    poolie_winners_this_season: List[int] = list(points_this_week.loc[points_this_week['pool_points']==points_this_week['pool_points'].max()]['poolteam_id'].values)
                    amount_won_per_winner = int((self.fee * len(points_this_week.index)) - points_by_week_pivot['Total Won'].sum() / len(poolie_winners_this_season))
                    for idx in points_by_week_pivot.index:
                        poolteam_id = points_by_week_pivot.loc[idx, 'poolteam_id']
                        if poolteam_id in poolie_winners_this_season:
                            kwargs = {'Criteria': [['id', '==', poolteam_id], ['pool_id', '==', self.id]]}
                            poolTeam = PoolTeam().fetch(**kwargs)
                            points_by_week_pivot.loc[idx, 'Total Won'] = points_by_week_pivot.loc[idx, 'Total Won'] + amount_won_per_winner
                            # update pool teams with amount won
                            poolTeam.won = points_by_week_pivot.loc[idx, 'Total Won']
                            poolTeam.persist(connection=connection)

                # display message showing pivoted table data
                old_stdout = sys.stdout
                new_stdout = io.StringIO()
                sys.stdout = new_stdout
                # don't want to show all weekly columns due to space restrictions
                data = points_by_week_pivot.copy(deep=True)
                if int(week_number) > 5:
                    col_list = list(data.columns.values)
                    cols_to_hide = []
                    start_week = int(week_number) - 4
                    for col in col_list:
                        if col.startswith('$'):
                            test_col = int(col.replace('$', ''))
                        elif col.isdecimal() is True:
                            test_col = int(col)
                        else:
                            continue
                        if test_col < start_week:
                            cols_to_hide.append(col)
                    data.drop(columns=cols_to_hide, inplace=True)
                print(data)
                output = new_stdout.getvalue()
                sys.stdout = old_stdout
                sg.PopupOK(output, title='Get Pool Standings', modal=True, line_width=120)

                # Save points_by_week_pivot to database
                points_by_week_pivot.to_sql('PoolTeamPointsByWeekPivot', con=connection, if_exists='replace', index=False)

        except Exception as e:
            sg.popup_error(f'Error in {sys._getframe().f_code.co_name}: {e}')
            raise

        return

    def get_pool_teams_as_df(self, pool: 'HockeyPool') -> pd.DataFrame:

        try:

            table_columns = [x['table column'] for x in pool_team_config['columns'] if 'table column' in x]

            sql = f'select {", ".join(table_columns)} from PoolTeam pt where pt.pool_id={pool.id}'

            if 'order by' in pool_team_config:
                order_by_criteria = [x for x in pool_team_config['order by']]
                if order_by_criteria:
                    sql = f'{sql} order by'
                    for d in order_by_criteria:
                        sql = f"{sql} {d['column']} {d['sequence']}"

            pool_teams = []
            with get_db_connection() as connection:
                cursor = connection.cursor()
                cursor.execute(sql)
                rows = cursor.fetchall()
                for row in rows:
                    pool_team = PoolTeam()
                    for key in row.keys():
                        value = row[key]
                        setattr(pool_team, key, value)
                    pool_teams.append(pool_team)

            cursor.close()
            connection.close()

            if len(pool_teams) == 0:
                pool_teams.append(PoolTeam())

            df = pd.DataFrame([pt.__dict__ for pt in pool_teams])

            runtime_columns = [x['runtime column'] for x in pool_team_config['columns'] if 'runtime column' in x]

        except Exception as e:
            sg.popup_error(f'Error in {sys._getframe().f_code.co_name}: {e}')

        return df

    def get_pool_teams_as_objects(self) -> List:

        # Get dataframe of pool teams
        connection = get_db_connection()
        sql = 'select id from PoolTeam where pool_id=? order by name'

        cursor = connection.cursor()
        cursor.execute(sql, (self.id,))

        pool_teams = []
        for row in cursor.fetchall():
            pool_team = PoolTeam()

            kwargs = {'Criteria': [['id', '==', row['id']]]}
            pool_team = PoolTeam().fetch(**kwargs)

            pool_teams.append(pool_team)

        connection.close()

        return pool_teams

    def getPool(self, id):

        self.fetch(**{'Criteria': [['id', '==', id]]})

        return self

    def getPoolTeamRostersByPeriod(self, batch: bool=False, pool_teams: List=[]):

        try:

            if batch:
                logger = logging.getLogger(__name__)
                dialog = None
            else:
                # layout the progress dialog
                layout = [
                    [
                        sg.Text(f'Get Pool Team Period Rosters for "{self.web_host}"...', size=(60,2), key='-PROG-')
                    ],
                    [
                        sg.Cancel()
                    ],
                ]
                # create the dialog
                dialog = sg.Window(f'Get Pool Team Period Rosters from "{self.web_host}"', layout, finalize=True, modal=True)

            fantrax = Fantrax(pool_id=self.id, league_id=self.league_id, season_id=self.season_id)
            dfPoolTeamPeriodRosters = fantrax.scrapePoolTeamPeriodRosters(season_id=self.season_id, pool_teams=pool_teams, dialog=dialog)
            del fantrax

            if len(dfPoolTeamPeriodRosters.index) == 0:
                if batch:
                    logger.debug('No pool team period rosters found. Returning...')
                return

            msg = f'Period rosters collected. Writing to database...'
            if batch:
                logger.debug(msg)
            else:
                dialog['-PROG-'].update(msg)
                event, values = dialog.read(timeout=2)
                if event == 'Cancel' or event == sg.WIN_CLOSED:
                    return

            if batch:
                logger.debug(f'Writing pool team period rosters for "{self.web_host}" to database')

            # Insert/update rosters
            with get_db_connection() as connection:
                if len(pool_teams) == 0:
                    sql = f"delete from dfPoolTeamPeriodRosters where season={self.season_id}"
                    connection.execute(sql)
                else:
                    for pool_team in pool_teams:
                        sql = f'delete from dfPoolTeamPeriodRosters where season={self.season_id} and pool_team="{pool_team}"'
                        connection.execute(sql)
                dfPoolTeamPeriodRosters.to_sql('dfPoolTeamPeriodRosters', con=connection, index=False, if_exists='append')

        except Exception as e:
            if batch:
                logger.error(repr(e))
            else:
                sg.popup_error(f'Error in {sys._getframe().f_code.co_name}: {e}')

        finally:
            msg = 'Update of pool team rosters completed...'
            if batch:
                logger.debug(msg)
            else:
                dialog.close()
                sg.popup_notify(msg, title=sys._getframe().f_code.co_name)

        return

    def getPlayers(self):

        NHL_API().get_players(season=season)

        return

    def getPlayerStats(self):

        global df_player_stats

        # If season has not started, the dataframe will be empty
        if season.SEASON_HAS_STARTED is False:
            # df_player_stats = ps.get_player_stats(season=season, pool=self, pt_roster=True)
            # need to get basic player info, without statistics
            sql = textwrap.dedent(f'''\
                select tr.*, t.id as team_id, ps.*
                from TeamRosters tr
                    left outer join PlayerStats ps on ps.player_id = tr.player_id and ps.seasonID == {season.id}
                    left outer join Team t on t.abbr == tr.team_abbr
                where tr.seasonID == {season.id}
            ''')
            df_player_stats = pd.read_sql(sql=sql, con=get_db_connection())
            # Drop the duplicate columns from the resulting DataFrame
            df_player_stats = df_player_stats.loc[:, ~df_player_stats.columns.duplicated(keep='first')]
            # df_player_stats = ps.merge_with_current_players_info(season=season, pool=self, df_stats=df_player_stats)
            df_player_stats = get_player_data.merge_with_current_players_info(season_id=str(season.id), pool_id=str(self.id), df_stats=df_player_stats)
            # df_player_stats = ps.merge_with_current_players_info(season=season, pool=self, df_stats=df_player_stats)
        else: # season.SEASON_HAS_STARTED is True
            df_player_stats = ps.get_player_stats(season=season, pool=self)
            # if season.SEASON_HAS_STARTED is True and season.type == 'P':
            #     next_season_id = season.getNextSeasonID()
            #     next_season_pool = HockeyPool().fetch(**{'Criteria': [['season_id', '==', next_season_id]]})
            #     df_player_stats = ps.merge_with_current_players_info(season=season, pool=next_season_pool, df_stats=df_player_stats)
            # else:
            #     df_player_stats = ps.merge_with_current_players_info(season=season, pool=self, df_stats=df_player_stats)
            # df_player_stats = ps.merge_with_current_players_info(season=season, pool=self, df_stats=df_player_stats)
            df_player_stats = get_player_data.merge_with_current_players_info(season_id=str(season.id), pool_id=str(self.id), df_stats=df_player_stats)

        return

    def import_athletic_projected_stats(self, season: Season):

        season_id = str(season.id)

        excel_path = f'{DATA_INPUT_FOLDER}/Projections/{season.id}/Athletic/{season_id[:4]}-{season_id[-2:]}-Fantasy-Projections-Fantrax.xlsx'

        excel_columns = ('NAME', 'POS', 'TEAM', 'ADP', 'GP', 'TOI', 'G', 'A', 'SOG','PPP', 'BLK', 'HIT', 'PIM', 'GP.1', 'W', 'SV', 'SV%', 'GAA')

        dfDraftList = pd.read_excel(io=excel_path, sheet_name='The List', header=0, usecols=excel_columns, engine='openpyxl')

        # rename columns
        dfDraftList.rename(columns={'NAME': 'Player', 'TEAM': 'Team', 'POS': 'Pos', 'GP': 'Games', 'G': 'Goals', 'A': 'Assists', 'HIT': 'Hits', 'BLK': 'BLKS', 'W': 'Wins', 'SV': 'Saves', 'SV%': 'Save%'}, inplace=True)

        # drop empty rows
        dfDraftList.dropna(subset=['Player'], how='any', inplace=True)

        # some player names are suffixed with " (C)", " (D)", " (G)"
        dfDraftList['Player'] = dfDraftList['Player'].apply(lambda x: x.replace(' (C)', '')
                                                                       .replace(' (D)', '')
                                                                       .replace(' (G)', '')
                                                           )

        # fix team abbr
        dfDraftList['Team'] = dfDraftList['Team'].apply(lambda x: x.replace('L.A', 'LAK')
                                                                   .replace('N.J', 'NJD')
                                                                   .replace('S.J', 'SJS')
                                                                   .replace('T.B', 'TBL')
                                                       )

        # fix 'F' position
        dfDraftList['Pos'] = dfDraftList['Pos'].apply(lambda x: x if x in ('C', 'LW', 'RW', 'D','G') else x.split('/')[0])

        # Convert ADP to float
        dfDraftList['ADP'] = pd.to_numeric(dfDraftList['ADP'], errors='coerce')

        # add season column
        dfDraftList['season_id'] = season.id

        # add player_id
        dfDraftList['player_id'] = assign_player_ids(df=dfDraftList, player_name='Player', nhl_team='Team', pos_code='Pos')

        ##############################################################################
        # Skaters
        ##############################################################################

        dfDraftList_skaters = dfDraftList.query('Pos!="G"').copy(deep=True)

        # drop not needed columns
        dfDraftList_skaters.drop(columns=['GP.1', 'Wins', 'Saves', 'Save%', 'GAA'], inplace=True)

        # set data types
        dtypes = {'Games': np.int32, 'TOI': np.int32, 'Goals': np.int32, 'Assists': np.int32, 'PPP': np.int32, 'PIM': np.int32, 'Hits': np.int32, 'BLKS': np.int32, 'SOG': np.int32}
        dfDraftList_skaters = dfDraftList_skaters.astype(dtypes)

        ##############################################################################
        # Goalies
        ##############################################################################

        dfDraftList_goalies = dfDraftList.query('Pos=="G"').copy(deep=True)

        # drop not needed columns
        dfDraftList_goalies.drop(columns=['Games', 'TOI', 'Goals', 'Assists', 'SOG', 'PPP', 'BLKS', 'Hits', 'PIM'], inplace=True)

        # rename columns
        dfDraftList_goalies.rename(columns={'GP.1': 'Games'}, inplace=True)

        # set data types
        dtypes = {'Games': np.int32, 'Wins': np.int32, 'Saves': np.int32}
        dfDraftList_goalies = dfDraftList_goalies.astype(dtypes)

        # GAA & Save%
        dfDraftList_goalies['GAA'] = dfDraftList_goalies['GAA'].round(2)
        dfDraftList_goalies['Save%'] = dfDraftList_goalies['Save%'].round(3)

        ##############################################################################
        # save to database
        ##############################################################################

        # remove duplicate player_id, if there are any. (i.e., Alex Vlasic)
        dfDraftList_skaters = dfDraftList_skaters[~dfDraftList_skaters.duplicated(subset=['player_id'], keep='last')]
        dfDraftList_goalies = dfDraftList_goalies[~dfDraftList_goalies.duplicated(subset=['player_id'], keep='last')]

        dfDraftList_skaters.to_sql('AthleticSkatersDraftList', con=get_db_connection(), index=False, if_exists='replace')
        dfDraftList_goalies.to_sql('AthleticGoaliesDraftList', con=get_db_connection(), index=False, if_exists='replace')

        return

    def import_dobber_projected_stats(self, season: Season):

        season_id = str(season.id)

        excel_path = f'{DATA_INPUT_FOLDER}/Projections/{season.id}/Dobber/dobberhockeydraftlist{season_id[:4]}{season_id[-2:]}.xlsx'

        ##############################################################################
        # Skaters
        ##############################################################################

        excel_columns = ('Rank', 'Player', 'Pos', '3YP', 'Upside', 'Team', 'Games', 'Goals', 'Assists', 'PP Unit', 'PIM', 'Hits', 'BLKS', 'SOG', 'Sleeper', 'Band-Aid Boy')

        dfDraftList_skaters = pd.read_excel(io=excel_path, sheet_name='EVERYTHING (Skaters)', header=5, usecols=excel_columns, engine='openpyxl')

        # drop empty rows
        dfDraftList_skaters.dropna(subset=['Player'], how='any', inplace=True)

        # add season column
        dfDraftList_skaters['season_id'] = season.id

        # Sebastian Aho (d)
        dfDraftList_skaters['Player'] = dfDraftList_skaters['Player'].apply(lambda x: x.replace(' (d)', ''))

        # change 'MON' team abbr to 'MTL'
        # NOTE: Dobber also has 'UFA' for unrestricted free agents.
        dfDraftList_skaters['Team'] = dfDraftList_skaters['Team'].apply(lambda x: 'MTL' if x == 'MON' else x)

        # add player_id
        dfDraftList_skaters['player_id'] = assign_player_ids(df=dfDraftList_skaters, player_name='Player', nhl_team='Team', pos_code='Pos')

        # convert NaN
        cols = {'Games': 0, 'Goals': 0, 'Assists': 0, 'Points': 0, 'PP Unit': 0, 'PIM': 0, 'Hits': 0, 'BLKS': 0, 'SOG': 0, 'Upside': 0, '3YP': 0}
        dfDraftList_skaters.fillna(cols, inplace=True)

        # set data types
        dtypes = {'Games': np.int32, 'Goals': np.int32, 'Assists': np.int32, 'PP Unit': np.int32, 'PIM': np.int32, 'Hits': np.int32, 'BLKS': np.int32, 'SOG': np.int32, 'Upside': np.int32, '3YP': np.int32}
        dfDraftList_skaters = dfDraftList_skaters.astype(dtypes)

        ##############################################################################
        # Goalies
        ##############################################################################

        excel_columns = ('Rank', 'Player', 'Team', 'Proj. Games', 'Wins', 'GAA', 'Band-Aid Boy', 'Notes')

        dfDraftList_goalies = pd.read_excel(io=excel_path, sheet_name='Goaltenders', header=5, usecols=excel_columns, engine='openpyxl')

        # drop empty rows
        dfDraftList_goalies.dropna(subset=['Player'], how='any', inplace=True)

        # rename columns
        dfDraftList_goalies.rename(columns={'Proj. Games': 'Games'}, inplace=True)

        # Need to add 'Pos' column for goalies
        dfDraftList_goalies['Pos'] = 'G'

       # add season column
        dfDraftList_goalies['season_id'] = season.id

        # change 'MON' team abbr to 'MTL'
        # NOTE: Dobber also has 'UFA' for unrestricted free agents.
        dfDraftList_goalies['Team'] = dfDraftList_goalies['Team'].apply(lambda x: 'MTL' if x == 'MON' else x)

        # add player_id
        dfDraftList_goalies['player_id'] = assign_player_ids(df=dfDraftList_goalies, player_name='Player', nhl_team='Team', pos_code='Pos')

        # convert NaN
        cols = {'Games': '', 'Wins': '', 'GAA': ''}
        dfDraftList_goalies.fillna(cols, inplace=True)

        # set data types
        dtypes = {'Games': np.int32, 'Wins': np.int32}
        dfDraftList_goalies = dfDraftList_goalies.astype(dtypes)

        ##############################################################################
        # save to database
        ##############################################################################

        # remove duplicate player_id, if there are any. (i.e., Alex Vlasic)
        dfDraftList_skaters = dfDraftList_skaters[~dfDraftList_skaters.duplicated(subset=['player_id'], keep='last')]
        dfDraftList_goalies = dfDraftList_goalies[~dfDraftList_goalies.duplicated(subset=['player_id'], keep='last')]

        dfDraftList_skaters.to_sql('DobberSkatersDraftList', con=get_db_connection(), index=False, if_exists='replace')

        dfDraftList_goalies.to_sql('DobberGoaliesDraftList', con=get_db_connection(), index=False, if_exists='replace')

        return

    def import_draft_picks(self, season: Season):

        dfDraftResults = pd.read_csv(f'{DATA_INPUT_FOLDER}/fantrax/{season.id}/Fantrax-Draft-Results-Vikings Fantasy Hockey League.csv', header=0)

        # drop rows with no player name
        dfDraftResults = dfDraftResults.dropna(subset=['Player'])

        # for idx, row in dfDraftResults.iterrows():
        #     # get player
        #     player_name = row['Player']
        #     if row['Team'] != '(N/A)':
        #         team_id = Team().get_team_id_from_team_abbr(team_abbr=row['Team'])
        #         # kwargs = {'full_name': player_name}
        #         # player = Player().fetch(**kwargs)
        #         # if player.id == 0:
        #         #     player_json = NHL_API().get_player_by_name(name=player_name, team_id=team_id)
        #         #     if player_json is None:
        #         #         poolie = row['Fantasy Team']
        #         #         msg = f'Player "{player_name}" not found for "{poolie}" pool team.'
        #         #         sg.popup_error(msg, title='import_draft_picks()')
        #         #         continue
        #         #     else:
        #         #         player.id = player_json['id']

        #     else:
        #         # In the 20232024 season, Patrick Kane is the only drafted player with team_abbr = '(N/A)'
        #         team_id = 0

        #     # get player id
        #     kwargs = get_player_id_from_name(name=player_name, team_id=team_id, pos='')
        #     player = Player().fetch(**kwargs)
        #     if player.id == 0:
        #         player_json = NHL_API().get_player_by_name(name=player_name, team_id=team_id)
        #         if player_json:
        #             player.id = player_json['id']
        #             player.full_name = player_json['fullName']

        #     dfDraftResults.loc[idx, 'player_id'] = player.id

        # add fantrax_id column
        dfDraftResults['fantrax_id'] = dfDraftResults['Player ID'].str.strip('*')

        # drop not needed columns
        columns_to_drop = ['Player ID']
        if 'Time (EDT)' in dfDraftResults.columns:
            columns_to_drop.append('Time (EDT)')
        else:
            columns_to_drop.append('Time (EST)')
        dfDraftResults.drop(columns=columns_to_drop, inplace=True)

        # add season column
        dfDraftResults['season_id'] = season.id

        # rename Player ID'
        dfDraftResults.rename(columns={'Player ID': 'fantrax_id'}, inplace=True)

        # add player_id
        dfDraftResults['player_id'] = assign_player_ids(df=dfDraftResults, player_name='Player', nhl_team='Team', pos_code='Pos', fantrax_id='fantrax_id')

        # rename columns
        dfDraftResults.rename(columns={'Round': 'round', 'Pick': 'pick', 'Ov Pick': 'overall', 'Player': 'player', 'Pos': 'pos', 'Team': 'team_abbr', 'Fantasy Team': 'pool_team'}, inplace=True)

        # player_id dtype to int
        dfDraftResults['player_id'] = dfDraftResults['player_id'].astype(int)

        # reindex columns
        dfDraftResults = dfDraftResults.reindex(columns=['season_id', 'round', 'pick', 'overall', 'player_id', 'player', 'pos', 'team_abbr', 'pool_team'])

        # delete season rows
        with get_db_connection() as connection:
            sql = f"delete from dfDraftResults where season_id={self.season_id}"
            connection.execute(sql)

        dfDraftResults.to_sql('dfDraftResults', con=get_db_connection(), index=False, if_exists='append')

        return

    def import_fantrax_projected_stats(self, season: Season):

        ##############################################################################
        # Skaters
        ##############################################################################

        excel_path = f'{DATA_INPUT_FOLDER}/Projections/{season.id}/Fantrax/Fantrax-Skaters.xls'

        excel_columns = ('ID', 'Player', 'Team', 'Position', 'Score', 'ADP', 'GP', 'G', 'A', 'PIM', 'SOG', 'PPP', 'Hit', 'Blk', 'Tk')

        dfDraftList_skaters = pd.read_excel(io=excel_path, sheet_name='Fantrax-Skaters', header=0, usecols=excel_columns, engine='xlrd')

        # remove players with 0 projected games
        dfDraftList_skaters.query('GP>0', inplace=True)

        # strip `*` from ID
        dfDraftList_skaters['ID'] = dfDraftList_skaters['ID'].str.strip('*')

        # rename columns
        dfDraftList_skaters.rename(columns={'Position': 'Pos', 'GP': 'Games', 'G': 'Goals', 'A': 'Assists', 'Hit': 'Hits', 'Blk': 'BLKS'}, inplace=True)

        # add season column
        dfDraftList_skaters['season_id'] = season.id

        # add player_id
        dfDraftList_skaters['player_id'] = assign_player_ids(df=dfDraftList_skaters, player_name='Player', nhl_team='Team', pos_code='Pos', fantrax_id='ID')

        ##############################################################################
        # Goalies
        ##############################################################################

        excel_path = f'{DATA_INPUT_FOLDER}/Projections/{season.id}/Fantrax/Fantrax-Goalies.xls'

        excel_columns = ('ID', 'Player', 'Team', 'Position', 'Score', 'ADP', 'GP', 'W', 'GAA', 'SV', 'SV%')

        dfDraftList_goalies = pd.read_excel(io=excel_path, sheet_name='Fantrax-Goalies', header=0, usecols=excel_columns, engine='xlrd')

        # remove players with 0 projected games
        dfDraftList_goalies.query('GP>0', inplace=True)

        # strip `*` from ID
        dfDraftList_goalies['ID'] = dfDraftList_goalies['ID'].str.strip('*')

        # rename columns
        dfDraftList_goalies.rename(columns={'Position': 'Pos', 'GP': 'Games', 'W': 'Wins', 'SV': 'Saves', 'SV%': 'Save%'}, inplace=True)

        # add season column
        dfDraftList_goalies['season_id'] = season.id

        # add player_id
        dfDraftList_goalies['player_id'] = assign_player_ids(df=dfDraftList_goalies, player_name='Player', nhl_team='Team', pos_code='Pos', fantrax_id='ID')

        # convert NaN
        cols = {'Wins': 0}
        dfDraftList_goalies.fillna(cols, inplace=True)

        # set data types
        dtypes = {'Wins': np.int32}
        dfDraftList_goalies = dfDraftList_goalies.astype(dtypes)

        ##############################################################################
        # save to database
        ##############################################################################

        # remove duplicate player_id, if there are any. (i.e., Alex Vlasic)
        dfDraftList_skaters = dfDraftList_skaters[~dfDraftList_skaters.duplicated(subset=['player_id'], keep='last')]
        dfDraftList_goalies = dfDraftList_goalies[~dfDraftList_goalies.duplicated(subset=['player_id'], keep='last')]

        dfDraftList_skaters.to_sql('FantraxSkatersDraftList', con=get_db_connection(), index=False, if_exists='replace')

        dfDraftList_goalies.to_sql('FantraxGoaliesDraftList', con=get_db_connection(), index=False, if_exists='replace')

        return

    def import_transaction_history(self, season: Season):

        season_id = str(season.id)

        dfHistoryClaimsDrops = pd.read_csv(f'{DATA_INPUT_FOLDER}/fantrax/{season_id}/Fantrax-Transaction-History-Claims+Drops-Vikings Fantasy Hockey League.csv', header=0)

        # add season column
        dfHistoryClaimsDrops['Season'] = season_id

        # rename columns
        dfHistoryClaimsDrops.rename(columns={'Position': 'Pos', 'Date (EDT)': 'DateTime', 'Team.1': 'Manager'}, inplace=True)

        # reformat DateTime
        dfHistoryClaimsDrops['DateTime'] = pd.to_datetime(dfHistoryClaimsDrops['DateTime'], format='%a %b %d, %Y, %I:%M%p').dt.strftime('%Y-%m-%d %H:%M')

        # Remove '<b>' and '</b>' tags from the 'Pos' column
        dfHistoryClaimsDrops['Pos'] = dfHistoryClaimsDrops['Pos'].str.replace('<b>', '', regex=False).str.replace('</b>', '', regex=False)

        # Replace 'TarasenKOâ€™d' with "TarasenKO'd" in 'Player' and 'Manager' columns
        dfHistoryClaimsDrops['Player'] = dfHistoryClaimsDrops['Player'].replace("TarasenKOâ€™d", "TarasenKO'd")
        dfHistoryClaimsDrops['Manager'] = dfHistoryClaimsDrops['Manager'].replace("TarasenKOâ€™d", "TarasenKO'd")

        # drop not needed columns
        dfHistoryClaimsDrops.drop(columns=['Fee'], inplace=True)

        # add PlayerId
        # Assign PlayerId for other players
        dfHistoryClaimsDrops['PlayerId'] = assign_player_ids(df=dfHistoryClaimsDrops, player_name='Player', nhl_team='Team', pos_code='Pos')
        # Handle specific case for Matt Murray
        dfHistoryClaimsDrops.loc[dfHistoryClaimsDrops['Player'] == 'Matt Murray', 'PlayerId'] = 8476899

        # PlayerId dtype to int
        dfHistoryClaimsDrops['PlayerId'] = dfHistoryClaimsDrops['PlayerId'].astype(int)

        # reindex columns
        dfHistoryClaimsDrops = dfHistoryClaimsDrops.reindex(columns=['Season', 'Period', 'DateTime', 'Manager', 'Type', 'PlayerId', 'Player', 'Team', 'Pos'])

        ##########################################################################################################################

        dfHistoryTrades = pd.read_csv(f'{DATA_INPUT_FOLDER}/fantrax/{season_id}/Fantrax-Transaction-History-Trades-Vikings Fantasy Hockey League.csv', header=0)

        # add season column
        dfHistoryTrades['Season'] = season_id

        # rename columns
        dfHistoryTrades.rename(columns={'Position': 'Pos', 'Date (EDT)': 'DateTime'}, inplace=True)

        # Replace null values in the 'Period' column with 0 and convert the column to integers
        dfHistoryTrades['Period'] = dfHistoryTrades['Period'].fillna(0).astype(int)

        # Replace null values in the 'Team' 'Pos' columns with ''
        dfHistoryTrades['Team'] = dfHistoryTrades['Team'].fillna('')
        dfHistoryTrades['Pos'] = dfHistoryTrades['Pos'].fillna('')

        # reformat DateTime
        dfHistoryTrades['DateTime'] = pd.to_datetime(dfHistoryTrades['DateTime'], format='%a %b %d, %Y, %I:%M%p').dt.strftime('%Y-%m-%d %H:%M')

        # Add 'Trade #' column
        trade_number = 0
        trade_numbers = [None] * len(dfHistoryTrades)
        for idx, fee in enumerate(dfHistoryTrades['Fees']):
            if pd.notna(fee) and fee != '':
                trade_number += 1
            trade_numbers[idx] = trade_number
        dfHistoryTrades['Trade #'] = trade_numbers

        # Renumber 'Trade #' so the highest value becomes 1, and the lowest value becomes the max
        max_trade_number = dfHistoryTrades['Trade #'].max()
        dfHistoryTrades['Trade #'] = dfHistoryTrades['Trade #'].apply(lambda x: max_trade_number - x + 1)

        # Remove '<b>' and '</b>' tags from the 'Pos' column
        dfHistoryTrades['Pos'] = dfHistoryTrades['Pos'].str.replace('<b>', '', regex=False).str.replace('</b>', '', regex=False)

        # Replace 'TarasenKOâ€™d' with "TarasenKO'd" in 'Player' and 'Manager' columns
        dfHistoryTrades['Player'] = dfHistoryTrades['Player'].replace("TarasenKOâ€™d", "TarasenKO'd")
        dfHistoryTrades['From'] = dfHistoryTrades['From'].replace("TarasenKOâ€™d", "TarasenKO'd")
        dfHistoryTrades['To'] = dfHistoryTrades['To'].replace("TarasenKOâ€™d", "TarasenKO'd")

        # drop not needed columns
        dfHistoryTrades.drop(columns=['Fees'], inplace=True)

        # add PlayerId
        # Handle specific case for Matt Murray
        dfHistoryTrades['PlayerId'] = assign_player_ids(
            df=dfHistoryTrades[~dfHistoryTrades['Player'].str.contains('Draft Pick', na=False)],
            player_name='Player',
            nhl_team='Team',
            pos_code='Pos'
        )
        dfHistoryTrades.loc[dfHistoryTrades['Player'] == 'Matt Murray', 'PlayerId'] = 8476899

        # PlayerId dtype to int
        dfHistoryTrades['PlayerId'] = dfHistoryTrades['PlayerId'].fillna(0).astype(int)

        # reindex columns
        dfHistoryTrades = dfHistoryTrades.reindex(columns=['Season', 'Trade #', 'Period', 'DateTime', 'From', 'To', 'PlayerId', 'Player', 'Team', 'Pos'])

        # Filter rows with '(Drop)' in the 'To' column
        drop_rows = dfHistoryTrades[dfHistoryTrades['To'] == '(Drop)']

        # Prepare the rows to be added to dfHistoryClaimsDrops
        drop_rows = drop_rows[['Season', 'Period', 'DateTime', 'From', 'Player', 'Team', 'Pos']].copy()
        drop_rows.rename(columns={'From': 'Manager'}, inplace=True)
        drop_rows['Type'] = 'Drop'

        # Append the rows to dfHistoryClaimsDrops
        dfHistoryClaimsDrops = pd.concat([dfHistoryClaimsDrops, drop_rows], ignore_index=True)

        # Remove the rows with '(Drop)' from dfHistoryTrades
        dfHistoryTrades = dfHistoryTrades[dfHistoryTrades['To'] != '(Drop)']

        # delete season rows
        with get_db_connection() as connection:
            sql = f"delete from dfHistoryTrades where Season={season_id}"
            try:
                connection.execute(sql)
            except sqlite3.OperationalError as e:
                if "no such table: dfHistoryTrades" in str(e):
                    pass
                else:
                    raise
            sql = f"delete from dfHistoryClaimsDrops where Season={season_id}"
            try:
                connection.execute(sql)
            except sqlite3.OperationalError as e:
                if "no such table: dfHistoryClaimsDrops" in str(e):
                    pass
                else:
                    raise

        dfHistoryTrades.to_sql('dfHistoryTrades', con=get_db_connection(), index=False, if_exists='append')

        dfHistoryClaimsDrops.to_sql('dfHistoryClaimsDrops', con=get_db_connection(), index=False, if_exists='append')

        return

    def layout_tab_pool_teams(self):

        layout = []
        rows = []

        (headings, visible_columns) = self.config_pool_team(web_host='Fantrax')
        rows.append(
            sg.pin(sg.Column(layout=[
                [
                    sg.Table([], headings=headings, auto_size_columns=False, visible_column_map=visible_columns, max_col_width=100, num_rows=26, justification='center', font=("Helvetica", 12), key='__FT_PT_MCLB__', vertical_scroll_only=False, alternating_row_color='dark slate gray', selected_row_colors=('black', 'SteelBlue2'), bind_return_key=True, enable_click_events=True, right_click_menu=['menu',['Refresh Roster', '-', 'Get Period Rosters']]),
                ]
            ], visible=True, key='__FT_PT_MCLB_CNTNR__')
        , vertical_alignment='top'))

        (headings, visible_columns) = self.config_pool_team_roster(web_host='Fantrax')
        rows.append(
            sg.pin(sg.Column(layout=[
                [
                    sg.Table([], headings=headings, auto_size_columns=False, visible_column_map=visible_columns, max_col_width=100, key='__FT_PTR_MCLB__', num_rows=26, justification='center', vertical_scroll_only=False, font=("Helvetica", 12), alternating_row_color='dark slate gray', selected_row_colors=('black', 'SteelBlue2'), bind_return_key=True, enable_click_events=True, right_click_menu=['menu',['Mark as Keeper', 'Mark as Minor', 'Unmark as Keeper/Minor', '-', 'Player Bio', '-','Remove Player']]),
                ]
            ], visible=True, key='__FT_PTR_MCLB_CNTNR__')
        , vertical_alignment='top'))

        layout.append(rows)

        return layout

    def layout_window(self):

        seasons = season.fetch_many()
        # if season.type == 'P':
        #     next_season_id = season.getNextSeasonID()
        #     pools = self.fetch_many(**{'Criteria': [['season_id', '==', next_season_id]]})
        # else:
        #     pools = self.fetch_many(**{'Criteria': [['season_id', '==', season.id]]})
        pools = self.fetch_many(**{'Criteria': [['season_id', '==', season.id]]})
        if len(pools) > 0:
            default_pool = pools[0]['name']
        else:
            default_pool = ''

        layout = [
            [sg.Menu(self.menu_bar(), key='MENU_BAR')],
            [sg.Button('Left Mouse Button', key='LEFT_MOUSE_BUTTON', visible=False, enable_events=True)],
            [sg.Button('Right Mouse Button', key='RIGHT_MOUSE_BUTTON', visible=False, enable_events=True)],
            [
                [
                    sg.Text('Season:', size=(7,1)),
                    sg.Combo(values=[x['name'] for x in seasons], default_value=season.name, size=(25,1), enable_events=True, key='__SEASON_COMBO__', metadata=seasons),
                ],
                [
                    sg.Text('Pool:', size=(7,1)),
                    sg.Combo(values=[x['name'] for x in pools], default_value=default_pool, size=(25,1), enable_events=True, key='__POOL_COMBO__', metadata=pools),
                ],
            ],
            [sg.TabGroup(layout=
                [
                    [
                        sg.Tab('Pool Teams', self.layout_tab_pool_teams(), key='__POOL_TEAMS_TAB__'),
                    ]
                ], title_color='black', key='Tab')
            ],
        ]

        return layout

    def make_clickable(self, column: str, value: str, alt_value: str='') -> str:

        # https://www.fantrax.com/fantasy/league/nhcwgeytkoxo2wc7/players;reload=2;statusOrTeamFilter=ALL;searchName=Jacob%20Bryson

        link = value

        if column == 'player_name':
            href = f"https://www.fantrax.com/fantasy/league/{self.league_id}/players;reload=2;statusOrTeamFilter=ALL;searchName={value.replace(' ', '%20')}"
            # target _blank to open new window
            link =  f'<a target="_blank" href="{href}">{value}</a>'

        return link

    def menu_bar(self):

        menu = [['&File', ['Refresh', '-', 'E&xit']],
                ['&Admin',
                    [
                        'NHL API...',
                            [
                            'Get Season Info',
                            'Get Teams',
                            '-',
                            'Get Players',
                            '-',
                            'Get Player Stats',
                            ],
                        '-',
                        'Get MoneyPuck Data',
                        '-',
                        'Update Player Injuries',
                        'Update Player Lines',
                        'Email Starting Goalie Projections',
                       '-',
                        'Fantrax...',
                            [
                                'Update Pool Teams',
                                'Update Pool Team Rosters',
                                'Update Fantrax Player Info',
                                '-',
                                'Update Standings Gain/Loss',
                                '-',
                                'Update Pool Team Service Times',
                                'Update Full Team Player Scoring',
                                '-',
                                'Import Watch List',
                                '-',
                                'Import Draft Picks',
                                'Import Transaction History',
                                '-',
                                'Email NHL Team Transactions',
                                '-',
                                'Get Pool Team Rosters, by Period',
                                '-',
                                'Archive Keeper Lists',
                            ],
                        '-',
                        'Manager Game Pace',
                        'Position Statistics',
                        '-',
                        'Summarize Trades History',
                        'Summarize Claims History',
                        '-',
                        'Start Flask Server',
                        '-',
                        'Start Daily VFHL Scheduled Task',
                        # 'Start "Hourly VFHL Activities"',
                        '-',
                        'Projected Stats...',
                            [
                                'Athletic Import',
                                'Dobber Import',
                                'Fantrax Import',
                            ],
                    ]
                ]]

        return menu

    def persist(self):

        try:

            with get_db_connection() as connection:

                sql = f'update Season set season=?, start_date=?, end_date=?, weels=? where id = {season.id}'
                values = [window.ReturnValuesDictionary['Season_Name'], window.ReturnValuesDictionary['Season_Start'], window.ReturnValuesDictionary['Season_End'], window.ReturnValuesDictionary['Weeks_In_Season']]
                connection.execute(sql, tuple(values))

        except Exception as e:
            msg = ''.join(traceback.format_exception(type(e), value=e, tb=e.__traceback__))
            sg.popup_error(msg)

        connection.close()

        return

    def set_pool(self, event_values):

        global season

        pool_dropdown = window['__POOL_COMBO__']

        with get_db_connection() as connection:
            cursor = connection.cursor()
            index = [i for i,x in enumerate(pool_dropdown.metadata) if x['name']==event_values['__POOL_COMBO__']][0]
            hp_oid = pool_dropdown.metadata[index]['id']
            sql = f'select hp.* from HockeyPool hp join Season s on s.id = hp.season_id where hp.id = {hp_oid} order by s.start_date'
            cursor.execute(sql)
            row = cursor.fetchone()

        cursor.close()
        connection.close()

        if row: # There will be 0 or 1
            self.id = row['id']
            self.name = row['name']
            self.season_id = row['season_id']
            self.web_host = row['web_host']
            self.draft_rule_id = row['draft_rule_id']
            self.fee = row['fee']
            self.league_id = row['league_id']

        season = Season().getSeason(id=season.id)
        season.set_season_constants()

        # Populates df_pool_team_rosters as global
        self.getPlayerStats()

        # Set hockey pool controls
        pool_dropdown.update(value=self.name)

        return

    def set_season(self, event_values):

        global season

        season_dropdown = window['__SEASON_COMBO__']

        index = [i for i,v in enumerate(season_dropdown.metadata) if season_dropdown.metadata[i]['name']==event_values['__SEASON_COMBO__']][0]
        season = season.getSeason(id=season_dropdown.metadata[index]['id'], type=season_dropdown.metadata[index]['type'])
        season.set_season_constants()

        pools = HockeyPool().fetch_many(**{'Criteria': [['season_id', '==', season.id]]})

        # Set pool
        pool_dropdown = window['__POOL_COMBO__']

        pool_dropdown.update(values=[x['name'] for x in pools])
        pool_dropdown.metadata = pools
        if len(pools) > 0:
            pool_dropdown.update(value=pools[0]['name'])
        hp = HockeyPool()
        if len(pools) > 0:
            hp.id = pools[0]['id']
            hp.name = pools[0]['name']
            hp.season_id = pools[0]['season_id']
            hp.web_host = pools[0]['web_host']
            hp.draft_rule_id = pools[0]['draft_rule_id']
            hp.fee = pools[0]['fee']
            hp.league_id = pools[0]['league_id']
            window['__POOL_TEAMS_TAB__'].update(visible=True)
        else:
            hp.season_id = season.id
            window['__POOL_TEAMS_TAB__'].update(visible=False)

        return hp

    def setCSS_TableStyles(self):

        # Set CSS properties for 'table' tag elements in dataframe
        table_props = [
        ('border', '1px solid black')
        ]

        # Set CSS properties for 'th' tag elements in dataframe
        th_props = [
        ('background', 'rgb(242, 242, 242)'),
        ('border', '1px solid black'),
        ('padding', '5px')
        ]

        # Set CSS properties for 'tr:nth-child' tag elements in dataframe
        # NOTE: This doesn't alter the background in the email table
        # However, when the email html source is displayed in a browser, the background does alternate
        tr_nth_child_props = [
        ('background', 'rgb(253, 233, 217)'),
        ]

        # Set CSS properties for 'td' tag elements in dataframe
        td_props = [
        ('border', '1px solid black'),
        ('padding', '5px'),
        ]

        # Set table styles
        styles = [
        dict(selector="table", props=table_props),
        dict(selector="th", props=th_props),
        dict(selector="td", props=td_props),
        dict(selector="tr:nth-child(even)", props=tr_nth_child_props)
        ]

        return styles

    def show_stat_summary_tables(self, season: Season):

        # # Get teams to save in dictionary
        # df_teams = pd.read_sql(f'select team_id, games from TeamStats where seasonID={season.id} and game_type="R"', con=get_db_connection())
        # teams_dict = {x.team_id: {'games': x.games} for x in df_teams.itertuples()}

        df_game_stats = get_player_data.get_game_stats(season_or_date_radios='season', from_season_id=season.id, to_season_id=season.id, from_date='', to_date='', pool_id='', game_type='R', ewma_span=10)

        if df_player_stats is None or df_player_stats.empty:
            return

        # aggregate cumulative stats per player
        df_cumulative = get_player_data.aggregate_game_stats(df=df_game_stats, stat_type='Cumulative')

        # calc max, mean, and std
        get_player_data.calc_scoring_category_maximums(df=df_cumulative)
        get_player_data.calc_scoring_category_minimums(df=df_cumulative)
        get_player_data.calc_scoring_category_means(df=df_cumulative)
        get_player_data.calc_scoring_category_std_deviations(df=df_cumulative)
        df_cumulative = get_player_data.calc_z_scores(df=df_cumulative, positional_scoring=True, stat_type='Cumulative')

        cumulative_max_cat = get_player_data.max_cat.copy()
        cumulative_min_cat = get_player_data.min_cat.copy()
        cumulative_mean_cat = get_player_data.mean_cat.copy()
        cumulative_std_cat = get_player_data.std_cat.copy()

        # aggregate per game stats per player
        df_per_game = get_player_data.aggregate_game_stats(df=df_game_stats, stat_type='Per game')

        # calc max, mean, and std
        get_player_data.calc_scoring_category_maximums(df=df_per_game)
        get_player_data.calc_scoring_category_minimums(df=df_per_game)
        get_player_data.calc_scoring_category_means(df=df_per_game)
        get_player_data.calc_scoring_category_std_deviations(df=df_per_game)
        df_per_game = get_player_data.calc_z_scores(df=df_per_game, positional_scoring=True, stat_type='Per game')

        per_game_max_cat = get_player_data.max_cat.copy()
        per_game_min_cat = get_player_data.min_cat.copy()
        per_game_mean_cat = get_player_data.mean_cat.copy()
        per_game_std_cat = get_player_data.std_cat.copy()

        from_year, to_year = split_seasonID_into_component_years(season_id=season.id)
        caption = f'<b><u>Statistics for the {from_year}-{to_year} season</u><br/>'

        ps.show_stat_summary_tables(df_cumulative=df_cumulative, df_per_game=df_per_game, cumulative_max_cat=cumulative_max_cat, cumulative_min_cat=cumulative_min_cat, cumulative_mean_cat=cumulative_mean_cat, cumulative_std_cat=cumulative_std_cat, per_game_max_cat=per_game_max_cat, per_game_min_cat=per_game_min_cat, per_game_mean_cat=per_game_mean_cat, per_game_std_cat=per_game_std_cat, caption=caption)

        return

    def start_daily_vfhl_scheduled_task(self):

        task_path = "\\Hockey Pool"
        task_name = "Daily VFHL Activities"

        # get task info
        cmd = f'powershell.exe "Get-ScheduledTask -TaskName \\"{task_name}\\" | Get-ScheduledTaskInfo"'
        process = subprocess.Popen(cmd, shell=True, stdout=subprocess.PIPE)
        result = process.communicate()[0].decode()
        # Parse the result string to get the values of interest
        lines = result.splitlines()
        last_run_time = next(line.split(': ')[1] for line in lines if 'LastRunTime' in line)
        last_task_result = next(line.split(': ')[1] for line in lines if 'LastTaskResult' in line)
        next_run_time = next(line.split(': ')[1] for line in lines if 'NextRunTime' in line)

        # check if task already running
        cmd = f'powershell.exe "Get-ScheduledTask -TaskName \\"{task_name}\\" | Select-Object -Property State"'
        process = subprocess.Popen(cmd, shell=True, stdout=subprocess.PIPE)
        result = process.communicate()[0].decode()
        # Parse the result string to get the state of the task
        state = result.split('\n')[3].strip()
        if state == 'Running':
            # ask if task should be killed
            user_response  = sg.popup_ok_cancel('Task Info', f'Task is still running.\nLast Run Time: {last_run_time}\nLast Task Result: {last_task_result}\nNext Run Time: {next_run_time}\n Do you want to kill it?')
            if user_response == 'OK':
                # stop task
                cmd = f'powershell.exe "Stop-ScheduledTask -TaskPath \\"{task_path}\\" -TaskName \\"{task_name}\\""'
                process = subprocess.Popen(cmd, shell=True, stdout=subprocess.PIPE)
                result = process.communicate()[0].decode()
            else:
                return

        # Display the values in a PySimpleGUI dialog
        # user_response  = sg.popup_ok_cancel('Task Info', f'Last Run Time: {last_run_time}\nLast Task Result: {last_task_result}\nNext Run Time: {next_run_time}\nDo you want to continue?')
        user_response = 'OK'
        if user_response == 'OK':
            # run task
            cmd = f'powershell.exe "Start-ScheduledTask -TaskPath \\"{task_path}\\" -TaskName \\"{task_name}\\""'
            process = subprocess.run(cmd, shell=True, stdout=subprocess.PIPE, check=True)
            ... # for debugging with a breakpoint

        return

    def summarize_claims_history(self):

        try:

            # Connect to the database
            with get_db_connection() as connection:
                # Query to get the required data
                sql = """
                    SELECT
                        Season,
                        Period,
                        Manager,
                        Player,
                        Team,
                        Pos
                    FROM dfHistoryClaimsDrops
                    WHERE Type = 'Claim'
                """
                # Execute the query and fetch the results
                dfClaimsHistory = pd.read_sql(sql, con=connection)

                sql = """
                    SELECT hp.season_id as Season, pt.name as Manager
                    FROM HockeyPool hp
                         JOIN PoolTeam pt on pt.pool_id=hp.id
                """
                dfPoolTeams = pd.read_sql(sql, con=connection)

                # Ensure all Season & Manager combinations in dfPoolTeams are also in dfClaimsHistory
                # Ensure both 'Season' columns have the same data type
                dfPoolTeams['Season'] = dfPoolTeams['Season'].astype(str)
                dfClaimsHistory['Season'] = dfClaimsHistory['Season'].astype(str)

                # Replace `McCarty’s Turtles` with `McCarty's Turtles` in both DataFrames
                dfPoolTeams['Manager'] = dfPoolTeams['Manager'].replace("McCarty’s Turtles", "McCarty's Turtles")
                dfClaimsHistory['Manager'] = dfClaimsHistory['Manager'].replace("McCarty’s Turtles", "McCarty's Turtles")
                # Replace `McCarty’s Turtles` with `McCarty's Turtles` in both DataFrames
                dfPoolTeams['Manager'] = dfPoolTeams['Manager'].replace("TarasenKO’d", "TarasenKO'd")
                dfClaimsHistory['Manager'] = dfClaimsHistory['Manager'].replace("TarasenKO’d", "TarasenKO'd")

                # Ensure unique combinations of Season and Manager in dfClaimsHistory
                unique_claims_history = dfClaimsHistory[['Season', 'Manager']].drop_duplicates()

                missing_combinations = dfPoolTeams.merge(
                    unique_claims_history,
                    on=['Season', 'Manager'],
                    how='left',
                    indicator=True
                ).query('_merge == "left_only"')[['Season', 'Manager']]

                if not missing_combinations.empty:
                    # Add missing combinations to dfClaimsHistory
                    missing_combinations['Pos'] = None  # No trade number for these entries
                    dfClaimsHistory = pd.concat([dfClaimsHistory, missing_combinations], ignore_index=True)

            # Count distinct claims per manager and season
            dfClaimsHistory_pivot = (
                dfClaimsHistory.groupby(['Season', 'Manager'])
                .size()
                .reset_index(name='Count')
                .pivot(index='Season', columns='Manager', values='Count')
            )

            # Add a column for the average number of trades per manager for each season
            dfClaimsHistory_pivot['Average'] = dfClaimsHistory_pivot[dfClaimsHistory_pivot > 0].mean(axis=1).round(1)

            # total number of trades
            dfClaimsHistory_pivot['Total'] = dfClaimsHistory_pivot[dfClaimsHistory_pivot.columns.difference(['Season', 'Average'])].sum(axis=1)

            # Reset the index for better readability
            dfClaimsHistory_pivot.reset_index(inplace=True)

            season_managers = dfClaimsHistory['Manager'].unique()
            dfClaimsHistory_pivot = dfClaimsHistory_pivot[['Season'] + sorted(list(season_managers)) + ['Total', 'Average']]

            # Replace NaN values with an empty string
            dfClaimsHistory_pivot.fillna('', inplace=True)

            # Ensure count columns are integers, except for 'Average'
            for column in dfClaimsHistory_pivot.columns:
                # if column not in ['Average', 'Season', 'Past Mgrs Avg']:
                if column not in ['Average', 'Season']:
                    dfClaimsHistory_pivot[column] = dfClaimsHistory_pivot[column].apply(lambda x: int(x) if x != '' else '')

            # Write the summary table
            # Transpose the dataframe
            dfClaimsHistory_pivot = dfClaimsHistory_pivot.set_index('Season').transpose()

            # Save the transposed dataframe to the database
            dfClaimsHistory_pivot.to_sql('dfHistoryClaims_Summary', con=get_db_connection(), index=True, if_exists='replace')

        except Exception as e:
            sg.popup_error(f"Error in summarizing claims history: {e}")

        return

    def summarize_trades_history(self):

        try:

            # Connect to the database
            with get_db_connection() as connection:
                # Query to get the required data
                sql = """
                    SELECT
                        Season,
                        "Trade #",
                        "From" AS Manager
                    FROM dfHistoryTrades
                    WHERE "From" != '*Deleted*' AND "To" != '*Deleted*'
                    UNION ALL
                    SELECT
                        Season,
                        "Trade #",
                        "To" AS Manager
                    FROM dfHistoryTrades
                    WHERE "From" != '*Deleted*' AND "To" != '*Deleted*'
                """
                # Execute the query and fetch the results
                dfHistoryTrades = pd.read_sql(sql, con=connection)

                sql = """
                    SELECT hp.season_id as Season, pt.name as Manager
                    FROM HockeyPool hp
                         JOIN PoolTeam pt on pt.pool_id=hp.id
                """
                dfPoolTeams = pd.read_sql(sql, con=connection)

                # Ensure all Season & Manager combinations in dfPoolTeams are also in dfHistoryTrades

                # Ensure both 'Season' columns have the same data type
                dfPoolTeams['Season'] = dfPoolTeams['Season'].astype(str)
                dfHistoryTrades['Season'] = dfHistoryTrades['Season'].astype(str)

                # Replace `McCarty’s Turtles` with `McCarty's Turtles` in both DataFrames
                dfPoolTeams['Manager'] = dfPoolTeams['Manager'].replace("McCarty’s Turtles", "McCarty's Turtles")
                dfHistoryTrades['Manager'] = dfHistoryTrades['Manager'].replace("McCarty’s Turtles", "McCarty's Turtles")
                # Replace `McCarty’s Turtles` with `McCarty's Turtles` in both DataFrames
                dfPoolTeams['Manager'] = dfPoolTeams['Manager'].replace("TarasenKO’d", "TarasenKO'd")
                dfHistoryTrades['Manager'] = dfHistoryTrades['Manager'].replace("TarasenKO’d", "TarasenKO'd")

                # Ensure unique combinations of Season and Manager in dfHistoryTrades
                unique_trades_history = dfHistoryTrades[['Season', 'Manager']].drop_duplicates()

                missing_combinations = dfPoolTeams.merge(
                    unique_trades_history,
                    on=['Season', 'Manager'],
                    how='left',
                    indicator=True
                ).query('_merge == "left_only"')[['Season', 'Manager']]

                if not missing_combinations.empty:
                    # Add missing combinations to dfHistoryTrades
                    missing_combinations['Trade #'] = None  # No trade number for these entries
                    dfHistoryTrades = pd.concat([dfHistoryTrades, missing_combinations], ignore_index=True)

            # Count distinct trades per manager and season
            dfHistoryTrades_Summary = (
                dfHistoryTrades.groupby(['Season', 'Manager'])
                .nunique()['Trade #']
                .reset_index()
                .pivot(index='Season', columns='Manager', values='Trade #')
            )

            # Add a column for the average number of trades per manager for each season
            dfHistoryTrades_Summary['Average'] = dfHistoryTrades_Summary[dfHistoryTrades_Summary > 0].mean(axis=1).round(1)

            # total number of trades
            dfHistoryTrades_Summary['Total'] = dfHistoryTrades.groupby('Season')['Trade #'].max() - dfHistoryTrades.groupby('Season')['Trade #'].min() + 1

            # Reset the index for better readability
            dfHistoryTrades_Summary.reset_index(inplace=True)

            season_managers = dfHistoryTrades['Manager'].unique()
            dfHistoryTrades_Summary = dfHistoryTrades_Summary[['Season'] + sorted(list(season_managers)) + ['Total', 'Average']]

            # Replace NaN values with an empty string
            dfHistoryTrades_Summary.fillna('', inplace=True)

            # Ensure count columns are integers, except for 'Average'
            for column in dfHistoryTrades_Summary.columns:
                # if column not in ['Average', 'Season', 'Past Mgrs Avg']:
                if column not in ['Average', 'Season']:
                    dfHistoryTrades_Summary[column] = dfHistoryTrades_Summary[column].apply(lambda x: int(x) if x != '' else '')

            # Write the summary table

            # Transpose the dataframe
            dfHistoryTrades_Summary = dfHistoryTrades_Summary.set_index('Season').transpose()

            # Save the transposed dataframe to the database
            dfHistoryTrades_Summary.to_sql('dfHistoryTrades_Summary', con=get_db_connection(), index=True, if_exists='replace')

            ####################################################################################################

            # cross-reference of trade counts for each manager

            # Step 1: Get Trades History
            with get_db_connection() as connection:
                # Query to get the required data
                sql = """
                    SELECT *
                    FROM dfHistoryTrades
                    WHERE [From] != "*Deleted*"
                """
                dfHistoryTrades = pd.read_sql(sql, con=connection)

            # Step 2: Get current season managers
            current_season = dfHistoryTrades['Season'].max()
            with get_db_connection() as connection:
                sql = f"""
                    SELECT hp.season_id as Season, pt.name as Manager
                    FROM HockeyPool hp
                         JOIN PoolTeam pt on pt.pool_id=hp.id
                    WHERE hp.season_id = {current_season}
                """
                dfCurrentManagers = pd.read_sql(sql, con=connection)

            current_managers = dfCurrentManagers['Manager']

            # Step 3: Filter dfHistoryTrades to include only trades involving current season managers
            dfFiltered = dfHistoryTrades[
                (dfHistoryTrades['From'].isin(current_managers)) &
                (dfHistoryTrades['To'].isin(current_managers))
            ]

            # Step 4: Create Manager1 and Manager2 columns
            # dfFiltered['Manager1'] = dfFiltered.apply(lambda row: min(row['From'], row['To']), axis=1)
            # dfFiltered['Manager2'] = dfFiltered.apply(lambda row: max(row['From'], row['To']), axis=1)
            dfFiltered.rename(columns={'From': 'Manager1', 'To': 'Manager2'}, inplace=True)

            # Ensure all Season & Manager combinations in dfPoolTeams are also in dfHistoryTrades
            # # Ensure both 'Season' columns have the same data type
            dfCurrentManagers['Season'] = dfCurrentManagers['Season'].astype(str)
            dfFiltered['Season'] = dfFiltered['Season'].astype(str)

            # Ensure unique combinations of Season and Manager in dfHistoryTrades
            unique_manager1_history = dfFiltered[['Season', 'Manager1']].drop_duplicates().rename(columns={'Manager1': 'Manager'})

            missing_combinations = dfCurrentManagers.merge(
                unique_manager1_history,
                on=['Season', 'Manager'],
                how='left',
                indicator=True
            ).query('_merge == "left_only"')[['Season', 'Manager']]

            if not missing_combinations.empty:
                # Add missing combinations to dfFiltered
                missing_combinations['Trade #'] = None  # No trade number for these entries
                missing_combinations.rename(columns={'Manager': 'Manager1'}, inplace=True)
                for manager2 in dfFiltered['Manager2'].unique():
                    temp_combinations = missing_combinations.copy()
                    temp_combinations['Manager2'] = manager2
                    dfFiltered = pd.concat([dfFiltered, temp_combinations], ignore_index=True)

            unique_manager2_history = dfFiltered[['Season', 'Manager2']].drop_duplicates().rename(columns={'Manager2': 'Manager'})

            missing_combinations = dfCurrentManagers.merge(
                unique_manager2_history,
                on=['Season', 'Manager'],
                how='left',
                indicator=True
            ).query('_merge == "left_only"')[['Season', 'Manager']]

            if not missing_combinations.empty:
                # Add missing combinations to dfFiltered
                missing_combinations['Trade #'] = None  # No trade number for these entries
                missing_combinations.rename(columns={'Manager': 'Manager2'}, inplace=True)
                for manager1 in dfFiltered['Manager1'].unique():
                    temp_combinations = missing_combinations.copy()
                    temp_combinations['Manager1'] = manager1
                    dfFiltered = pd.concat([dfFiltered, temp_combinations], ignore_index=True)

            # Step 5: Group by Manager1 and Manager2 and count distinct 'Trade #' values
            dfGrouped = (
                dfFiltered.groupby(['Season', 'Manager1', 'Manager2'])['Trade #']
                .nunique()
                .reset_index(name='TradeCount')
            )

            # Step 6: Pivot the grouped DataFrame to create the cross-reference table
            dfHistoryTrades_XRef = dfGrouped.pivot_table(index='Manager1', columns='Manager2', values='TradeCount', aggfunc='sum', fill_value=0).astype(int)

            # Set diagonal values (where Manager1 == Manager2) to empty strings
            for manager in dfHistoryTrades_XRef.index:
                if manager in dfHistoryTrades_XRef.columns:
                    dfHistoryTrades_XRef.at[manager, manager] = ''

            # # Replace all 0 values with an empty string
            # dfHistoryTrades_XRef.replace(0, '', inplace=True)

            # Save the transposed dataframe to the database
            dfHistoryTrades_XRef.to_sql('dfHistoryTrades_XRef', con=get_db_connection(), index=True, if_exists='replace')

        except Exception as e:
            sg.popup_error(f"Error in summarizing trades history: {e}")

        return

    def table_container_formatting(self, config_columns, mclb):

        # get character width for character that is closest to average character width
        char_width = sg.Text.char_width_in_pixels(mclb.Font, character='n')
        # get maximum width for each column
        col_max_widths = [x['max width'] if 'max width' in x else None for x in config_columns]
        # get column justifications
        col_anchors = ['w' if ('justify' in x and x['justify']=='left') else 'e' if ('justify' in x and x['justify']=='left') else 'center' for x in config_columns]
        # get column formatting string
        col_formats = [x['format'] if 'format' in x else None  for x in config_columns]
        # get visible columns
        visible_columns = [x['title'] for x in config_columns if 'visible' not in x or ('visible' in x and x['visible'] is True)]
        # get column headings
        headings = [x['title'] for x in config_columns]

        # Expand table in both directions of 'x' and 'y'
        mclb.expand(expand_x=True, expand_y=True)
        # for cid in visible_columns:
        #     # Set column stretchable when window resize
        #     mclb.Widget.column(cid, stretch=True)

        # get mclb data
        data = [[x for x in row] for row in mclb.Values]
        # transpose data
        data = list(map(list, zip(*data)))
        # apply column formatting, as appropriate
        outer_list = []
        try:
            for format_str, column_values in zip(col_formats, data):
                inner_list = []
                for value in column_values:
                    # mclbs with "Totals" rows may have formatting, but no aggregated value (i.e., value=='')
                    if isfunction(format_str) is True:
                        # format_int & format_2_decimals lambda formatting funtions
                        inner_list.append(format_str(value))
                    elif format_str is None or pd.isna(value) or value=='nan' or value=='':
                        inner_list.append(value)
                    else: # formating numeric
                        if 'f' in format_str: # float
                            inner_list.append('' if float(value)==0.0 else format_str.format(float(value)))
                        else: # integer
                            inner_list.append('' if (int(value)==0 and format_str in ('{:,}', '{:}')) else format_str.format(int(value)))
                outer_list.append(inner_list)
        except Exception as e:
            msg = ''.join(traceback.format_exception(type(e), value=e, tb=e.__traceback__))
            sg.popup_error(msg)
        # transpose data back
        data = list(map(list, zip(*outer_list)))

        # All data to include headings when determining column width
        all_data = [headings] + data

        # Find width in pixel and 3 extra characters for each column
        col_widths = [min([max(map(len, map(str,cols)))+3, 999 if col_max_widths[i] is None else col_max_widths[i]])*char_width for i, cols in enumerate(zip(*all_data))]

        # update all new data, with formatting
        mclb.update(values=data)

        # Update table columns with new widths & justification anchors
        # NOTE: if horizontal scrollbar not used, watch for widget too large to fit window or screen.
        mclb.Widget.pack_forget()
        for cid, width, anchor in zip(headings, col_widths, col_anchors):
            # Set column stretchable when window resize
            mclb.Widget.column(cid, stretch=True)
            mclb.Widget.column(cid, width=width)
            mclb.Widget.column(cid, anchor=anchor)
        # Redraw table
        mclb.Widget.pack(side='left', fill='both', expand=True)

        return

    def updatePlayerInjuries(self, suppress_prompt: bool=False, batch: bool=False):

        try:

            logger = logging.getLogger(__name__)

            if suppress_prompt is True or batch is True:
                response = 1
            else:
                response = self.askToScrapeWebPage()

            if batch is True:
                dialog = None
            else:
                # progress dialog
                layout = [
                    [
                        sg.Text(f'Update Player Injuries...', size=(100, 3), key='-PROG-')
                    ],
                    [
                        sg.Cancel()
                    ],
                ]
                dialog = sg.Window(f'Update Player Injuries', layout, modal=True, finalize=True)


            if response == 1:
                dfPlayerInjuries = injuries.main(dialog=dialog)
                if not batch:
                    dialog['-PROG-'].update(f'Injuries collected for {len(dfPlayerInjuries.index)} players. Writing to database...')
                    event, values = dialog.read(timeout=2)
                    if event == 'Cancel' or event == sg.WIN_CLOSED:
                        return

                dfPlayerInjuries.to_sql('dfPlayerInjuries', sqlite3.connect(DATABASE), index=False, if_exists='replace')
            elif response == 2:
                # print('Getting player injuries from database')
                if not batch:
                    dialog['-PROG-'].update('Getting player injuries from database...')
                    event, values = dialog.read(timeout=2)
                    if event == 'Cancel' or event == sg.WIN_CLOSED:
                        return
                dfPlayerInjuries = pd.read_sql('select * from dfPlayerInjuries', sqlite3.connect(DATABASE))
            else:
                return

            with get_db_connection() as connection:
                # Before updating, clear current injury info
                sql = "update Player set injury_status = '', injury_note = ''"
                cursor = connection.cursor()
                cursor.execute(sql)
                connection.commit()

            # # Get team IDs, player names & id dictionary, and NHL_API
            # team_ids = load_nhl_team_abbr_and_id_dict()
            # player_ids = load_player_name_and_id_dict()
            # nhl_api = NHL_API()

            for idx in dfPlayerInjuries.index:

                # Update Player
                playerName = dfPlayerInjuries['name'][idx]
                playerName = playerName.strip()
                playerID = dfPlayerInjuries['id'][idx]
                # if playerID == 0 and playerName == 'Elias Pettersson':
                #     playerID = 8480012 # Elias Pettersson: Forward
                #     # playerID = 8483678 # Elias Pettersson: Defenseman
                playerPos = dfPlayerInjuries['pos'][idx]
                playerTeam = dfPlayerInjuries['team'][idx]
                injuryStatus = ''.join([dfPlayerInjuries['status'][idx], ' (', dfPlayerInjuries['date'][idx], ')'])
                injuryNote = dfPlayerInjuries['note'][idx]

                # Get NHL Player
                # get team_id from team_abbr
                # team_id = Team().get_team_id_from_team_abbr(team_abbr=playerTeam, suppress_error=True)
                # kwargs = get_player_id_from_name(name=playerName, team_id=team_id)
                # nhlPlayer = Player().fetch(**kwargs)
                # if nhlPlayer.id == 0:
                #     player_json = NHL_API().get_player_by_name(name=playerName, team_id=team_id)
                #     if player_json is None:
                #         msg = f'There are no NHL players with name "{playerName}".'
                #         if batch:
                #             logger.error(msg)
                #         else:
                #             sg.popup_ok(msg)
                #         continue

                # team_id = team_ids[playerTeam] if playerTeam != 'N/A' else 0
                # player_id = get_player_id(team_ids, player_ids, nhl_api, playerName, playerTeam, playerPos)
                if playerID == 0:
                    msg = f'updatePlayerInjuries(): There are no NHL players with name "{playerName}".'
                    if batch:
                        logger.error(msg)
                    else:
                        sg.popup_ok(msg)
                    continue

                kwargs = {'id': playerID}
                nhlPlayer = Player().fetch(**kwargs)

                nhlPlayer.injury_status = injuryStatus
                nhlPlayer.injury_note = injuryNote
                if nhlPlayer.persist() is False:
                    # print(f'Could not perist NHL player "{nhlPlayer.full_name}"')
                    msg = f'Could not perist NHL player "{nhlPlayer.full_name}"'
                    if batch:
                        logger.debug(msg)
                    else:
                        sg.popup_ok(msg)


        except Exception as e:
            if batch:
                logger.error(repr(e))
                raise
            else:
                sg.PopupScrolled(''.join(traceback.format_exception(type(e), value=e, tb=e.__traceback__)), modal=True)

        finally:
            msg = 'Update of player injuries completed...'
            if batch:
                logger.debug(msg)
            else:
                dialog.close()
                sg.popup_notify(msg, title=sys._getframe().f_code.co_name)

        return

    def update_player_lines(self, season: Season, batch: bool=False, game_date: str=date.strftime(date.today(), '%Y-%m-%d')):

        try:

            logger = logging.getLogger(__name__)

            if batch is True:
                dialog = None
            else:
                # progress dialog
                layout = [
                    [
                        sg.Text(f'Update Player Lines...', size=(100, 3), key='-PROG-')
                    ],
                    [
                        sg.Cancel()
                    ],
                ]
                dialog = sg.Window(f'Update Player Lines', layout, modal=True, finalize=True)

            if batch:
                # logger.debug(f'Calling player_lines.from_daily_fantasy_fuel() with dialog={dialog}, game_date={game_date})')
                logger.debug('Calling player_lines.from_daily_faceoff()')

            # dfPlayerLines = player_lines.from_daily_fantasy_fuel(dialog=dialog, game_date=game_date)
            dfPlayerLines = player_lines.from_daily_faceoff(dialog=dialog, batch=batch)

            if batch:
                # logger.debug(f'Return from player_lines.from_daily_fantasy_fuel() with dfPlayerLines size={len(dfPlayerLines)}')
                logger.debug('Return from player_lines.from_daily_faceoff()')

            if len(dfPlayerLines.index) == 0:
                msg = 'No player lines collected. Returning...'
                if batch:
                    logger.error(msg)
                else:
                    dialog['-PROG-'].update(msg)
                    event, values = dialog.read(timeout=2)
                return
            else:
                msg = f'Player lines collected for {len(dfPlayerLines.index)} players. Writing to database...'
                if batch:
                    logger.debug(msg)
                else:
                    dialog['-PROG-'].update(msg)
                    event, values = dialog.read(timeout=2)
                    if event == 'Cancel' or event == sg.WIN_CLOSED:
                        return

            # add player_id
            dfPlayerLines['player_id'] = assign_player_ids(df=dfPlayerLines, player_name='name', nhl_team='team', pos_code='pos')

            # write to database
            dfPlayerLines.to_sql('dfPlayerLines', sqlite3.connect(DATABASE), index=False, if_exists='replace')

            # sort by team
            teams = dfPlayerLines['team'].unique()
            for team in teams:
                sql = f'update TeamRosters set line="", pp_line="" where seasonID={season.id} and team_abbr="{team}"'
                with get_db_connection() as connection:
                    connection.execute(sql)

            for idx in dfPlayerLines.index:

                # # bypass goalies
                # if dfPlayerLines['pos'][idx] == 'G':
                #     continue

                # Update Player
                player_id = dfPlayerLines['player_id'][idx]
                name = dfPlayerLines['name'][idx]
                # if name == 'Tim Stuetzle':
                #     name = 'Tim Stutzle'
                team = dfPlayerLines['team'][idx]
                pos = dfPlayerLines['pos'][idx]
                line = dfPlayerLines['line'][idx]
                pp_line = dfPlayerLines['pp_line'][idx]

                # # Get NHL Player
                # # get team_id from team_abbr
                # team_id = Team().get_team_id_from_team_abbr(team_abbr=team, suppress_error=True)
                # kwargs = get_player_id_from_name(name=name, team_id=team_id)
                # nhlPlayer = Player().fetch(**kwargs)
                # if nhlPlayer.id == 0:
                #     player_json = NHL_API().get_player_by_name(name=name, team_id=team_id)
                #     if player_json is None:
                #         msg = f'There are no NHL players with name "{name}".'
                #     else:
                #         msg = f'NHL player "{name}" (id={player_json["playerId"]}) not found in Player table.'
                #     if batch:
                #         logger.error(msg)
                #     else:
                #         sg.popup_ok(msg)
                #     continue
                # # else:
                # #     nhlPlayer = nhlPlayers[0]

                # update team roster player
                sql = f'''
                    INSERT OR REPLACE INTO TeamRosters (seasonID, player_id, team_abbr, name, pos, line, pp_line)
                    VALUES ({season.id}, {player_id}, "{team}", "{name}", "{pos}", "{line}", "{pp_line}");
                '''
                with get_db_connection() as connection:
                    try:
                        connection.execute(sql)
                    except Exception as e:
                        sg.popup_error(f'Error in {sys._getframe().f_code.co_name}: {e}')

        except Exception as e:
            if batch:
                logger.error(repr(e))
                raise
            else:
                sg.PopupScrolled(''.join(traceback.format_exception(type(e), value=e, tb=e.__traceback__)), modal=True)

        finally:
            msg = 'Update of player lines completed...'
            if batch:
                logger.debug(msg)
            else:
                dialog.close()
                sg.popup_notify(msg, title=sys._getframe().f_code.co_name)

        return

    def updateFantraxPlayerInfo(self, batch: bool=False, watchlist: bool=False):

        try:

            if batch:
                logger = logging.getLogger(__name__)
                dialog = None
            else:
                # layout the progress dialog
                layout = [
                    [
                        sg.Text(f'Update Player Info...', size=(60,2), key='-PROG-')
                    ],
                    [
                        sg.Cancel()
                    ],
                ]
                # create the dialog
                dialog = sg.Window(f'Fantrax Player Info', layout, finalize=True, modal=True)

            if self.web_host == 'Fantrax':
                fantrax = Fantrax(pool_id=self.id, league_id=self.league_id, season_id=self.season_id)
                dfFantraxPlayerInfo = fantrax.scrapePlayerInfo(dialog=dialog, watchlist=watchlist)
                del fantrax
            else:
                return

            if dfFantraxPlayerInfo.index is None:
                if batch:
                    logger.debug('Fantrax player info scrape failed. Returning...')
                return

            if len(dfFantraxPlayerInfo.index) == 0:
                if batch:
                    logger.debug('No player info found. Returning...')
                if watchlist is False:
                    return

            msg = f'Fantrax player info collected for {len(dfFantraxPlayerInfo.index)} players. Writing to database...'
            if batch:
                logger.debug(msg)
            else:
                dialog['-PROG-'].update(msg)
                event, values = dialog.read(timeout=2)
                if event == 'Cancel' or event == sg.WIN_CLOSED:
                    return

            # add season_id column (NOTE: season is global)
            dfFantraxPlayerInfo.insert(0, 'season_id', self.season_id)

            if watchlist is True:
                sql = f'update FantraxPlayerInfo set watch_list = 0 where season_id = {self.season_id}'
                with get_db_connection() as connection:
                    connection.execute(sql)
                    for row in dfFantraxPlayerInfo.itertuples():
                        if row.player_id and row.player_id > 0:
                            # minors = 1 if row.minors is True else 0
                            sql = dedent(f'''\
                                insert or replace into FantraxPlayerInfo
                                (season_id, player_id, fantrax_id, player_name, nhl_team, pos, minors, rookie, watch_list, score, next_opp)
                                values ({self.season_id}, {row.player_id}, "{row.fantrax_id}", "{row.player_name}", "{row.nhl_team}", "{row.pos}", {row.minors}, "{row.rookie}", 1, {row.score}, "{row.next_opp}")
                                ''')
                            connection.execute(sql)
                    connection.commit()
            else:
                # I tried dropping the table, but pandas to_sql creates the table without a primary key on the player_id,
                # which I need for watchlist processing (see above)
                sql = f'delete from FantraxPlayerInfo where season_id = {self.season_id}'
                with get_db_connection() as connection:
                    connection.execute(sql)
                    connection.commit()
                # remove duplicate player_id, if there are any.
                dfFantraxPlayerInfo = dfFantraxPlayerInfo[~dfFantraxPlayerInfo.duplicated(subset=['player_id'], keep='last')]
                dfFantraxPlayerInfo.to_sql('FantraxPlayerInfo', con=get_db_connection(), index=False, if_exists='append')

        except Exception as e:
            msg = ''.join(traceback.format_exception(type(e), value=e, tb=e.__traceback__))
            if batch:
                logger.error(msg)
                raise
            else:
                sg.popup_error(msg)

        finally:
            msg = 'Update of Fantrax player info completed...'
            if batch:
                logger.debug(msg)
            else:
                dialog.close()
                sg.popup_notify(msg, title=sys._getframe().f_code.co_name)

        return

    def updateFullTeamPlayerScoring(self, batch: bool=False):

        try:

            if batch:
                logger = logging.getLogger(__name__)
                dialog = None
            else:
                # layout the progress dialog
                layout = [
                    [
                        sg.Text(f'Update Full Team Player Scoring for "{self.web_host}"...', size=(60,2), key='-PROG-')
                    ],
                    [
                        sg.Cancel()
                    ],
                ]
                # create the dialog
                dialog = sg.Window(f'Update Full Team Player Scoring from "{self.web_host}"', layout, finalize=True, modal=True)
                logger = None

            fantrax = Fantrax(pool_id=self.id, league_id=self.league_id, season_id=self.season_id)
            dfFullTeamPlayerScoring = fantrax.scrapeFullTeamPlayerScoring(dialog=dialog)
            del fantrax

            if len(dfFullTeamPlayerScoring.index) == 0:
                if batch:
                    logger.debug('No team player scoring found. Returning...')
                return

            msg = f'Full Team Player Scoring collected. Updating dataframe...'
            if batch:
                logger.debug(msg)
            else:
                dialog['-PROG-'].update(msg)
                event, values = dialog.read(timeout=2)
                if event == 'Cancel' or event == sg.WIN_CLOSED:
                    return

            # add player id column
            dfFullTeamPlayerScoring['player_id'] = assign_player_ids(df=dfFullTeamPlayerScoring, player_name='player', nhl_team='team', pos_code='pos')

            #########################################################################################################
            # add pre-draft keeper columns
            sql = f'select player_id, case when keeper = "m" then "MIN" else "Yes" end as pre_draft_keeper, pool_team as pre_draft_manager from KeeperListsArchive where pool_id={self.id}'
            df_keeper_list = pd.read_sql(sql, con=get_db_connection())

            # add pre-draft keeper columns to df
            dfFullTeamPlayerScoring.set_index(['player_id'], inplace=True)
            df_keeper_list.set_index(['player_id'], inplace=True)

            dfFullTeamPlayerScoring['keeper'] = dfFullTeamPlayerScoring.apply(lambda row: 1 if row.name in df_keeper_list.index else 0, axis=1)

            dfFullTeamPlayerScoring['orig_keeper'] = dfFullTeamPlayerScoring.apply(lambda row: 1 if row.name in df_keeper_list.index and row['manager'] == df_keeper_list.loc[row.name, 'pre_draft_manager'] else 0, axis=1)

            dfFullTeamPlayerScoring['other_keeper'] = dfFullTeamPlayerScoring.apply(lambda row: 1 if row.name in df_keeper_list.index and row['manager'] != df_keeper_list.loc[row.name, 'pre_draft_manager'] else 0, axis=1)

            dfFullTeamPlayerScoring.reset_index(inplace=True)
            #########################################################################################################
            # Convert 'g_minutes' to 'g_seconds'
            dfFullTeamPlayerScoring['g_seconds'] = dfFullTeamPlayerScoring['g_minutes'].apply(lambda x: int(x.split(':')[0]) * 60 + int(x.split(':')[1]) if isinstance(x, str) and ':' in x else 0)

            # Reorder columns to position 'player_id' and 'current' after 'team'
            cols = dfFullTeamPlayerScoring.columns.tolist()

            player_id_index = cols.index('player_id')
            cols.insert(player_id_index + 1, cols.pop(cols.index('player_id')))

            status_index = cols.index('status')
            # and in reverse order to insert 'original' as the first column
            cols.insert(status_index + 1, cols.pop(cols.index('other_keeper')))
            cols.insert(status_index + 1, cols.pop(cols.index('orig_keeper')))
            cols.insert(status_index + 1, cols.pop(cols.index('keeper')))

            g_minutes_index = cols.index('g_minutes')
            cols.insert(g_minutes_index + 1, cols.pop(cols.index('g_seconds')))

            # set new column sequence
            dfFullTeamPlayerScoring = dfFullTeamPlayerScoring[cols]

            # drop 'g_minutes' column
            dfFullTeamPlayerScoring.drop(columns=['g_minutes'], inplace=True)

            # sort by manager and player columns
            dfFullTeamPlayerScoring.sort_values(by=['manager', 'player'], ascending=[True, True], inplace=True)

            msg = f'Writing to database...'
            if batch:
                logger.debug(msg)
            else:
                dialog['-PROG-'].update(msg)
                event, values = dialog.read(timeout=2)
                if event == 'Cancel' or event == sg.WIN_CLOSED:
                    return

            # write to database
            dfFullTeamPlayerScoring.to_sql('dfFullTeamPlayerScoring', con=get_db_connection(), index=False, if_exists='replace')

            msg = f'Writing dfFullTeamPlayerScoring to Excel...'
            if batch:
                logger.debug(msg)
            else:
                dialog['-PROG-'].update(msg)
                event, values = dialog.read(timeout=2)
                if event == 'Cancel' or event == sg.WIN_CLOSED:
                    return

            # write to Excel
            excelSheets = [
                {
                    'df': dfFullTeamPlayerScoring,
                    'excelSheet': 'Full Roster Scoring',
                    'min_row': 2,
                    'max_row': None,
                    'min_col': 1
                },
            ]
            # self.writeToExcel(df=dfFullTeamPlayerScoring, excelFile="Manager Player Service Times.xlsx", excelSheet='Full Roster Scoring', min_row=2, max_row=None, min_col=1, batch=batch, logger=logger, dialog=dialog)
            self.writeToExcel(excelFile="Manager Player Service Times.xlsx", excelSheets=excelSheets, batch=batch, logger=logger, dialog=dialog)

        except Exception as e:
            msg = f'Error in {sys._getframe().f_code.co_name}: {e}'
            if batch:
                logger.error(msg)
            else:
                sg.popup_error(msg)

        finally:
            msg = 'Update of Full Team Player Scoring completed...'
            if batch:
                logger.debug(msg)
            else:
                dialog.close()
                sg.popup_notify(msg, title=sys._getframe().f_code.co_name)

        return

    def updatePoolTeamRosters(self, suppress_prompt: bool=False, batch: bool=False, pool_teams: List=[]):

        try:

            if suppress_prompt is True:
                response = 1
            else:
                response = self.askToScrapeWebPage()

            if batch:
                logger = logging.getLogger(__name__)
                dialog = None
            else:
                # layout the progress dialog
                layout = [
                    [
                        sg.Text(f'Update Pool Team Rosters for "{self.web_host}"...', size=(60,2), key='-PROG-')
                    ],
                    [
                        sg.Cancel()
                    ],
                ]
                # create the dialog
                dialog = sg.Window(f'Update Pool Team Rosters from "{self.web_host}"', layout, finalize=True, modal=True)

            if response == 1:
                fantrax = Fantrax(pool_id=self.id, league_id=self.league_id, season_id=self.season_id)
                dfPoolTeamRoster = fantrax.scrapePoolTeamRosters(dialog=dialog, pool_teams=pool_teams)
                del fantrax
            elif response == 2:
                msg = 'Getting pool team rosters data from database...'
                if batch:
                    logger.debug(msg)
                else:
                    dialog['-PROG-'].update(msg)
                    event, values = dialog.read(timeout=2)
                    if event == 'Cancel' or event == sg.WIN_CLOSED:
                        return
                dfPoolTeamRoster = pd.read_sql('select * from dfPoolTeamPlayers',con=get_db_connection())
            else:
                return

            if dfPoolTeamRoster is None or len(dfPoolTeamRoster.index) == 0:
                if batch:
                    logger.debug('No pool team rosters found. Returning...')
                return

            msg = f'Rosters collected for {len(dfPoolTeamRoster.index)} players. Writing to database...'
            if batch:
                logger.debug(msg)
            else:
                dialog['-PROG-'].update(msg)
                event, values = dialog.read(timeout=2)
                if event == 'Cancel' or event == sg.WIN_CLOSED:
                    return

            if batch:
                logger.debug(f'Updating pool team rosters for "{self.web_host}"')
            # Insert/update rosters
            fantrax = Fantrax(pool_id=self.id, league_id=self.league_id, season_id=self.season_id)
            fantrax.updatePoolTeamRosters(pool=self, df=dfPoolTeamRoster, pool_teams=pool_teams, batch=True)

        except Exception as e:
            if batch:
                logger.error(repr(e))
            else:
                sg.popup_error(f'Error in {sys._getframe().f_code.co_name}: {e}')

        finally:
            msg = 'Update of pool team rosters completed...'
            if batch:
                logger.debug(msg)
            else:
                dialog.close()
                sg.popup_notify(msg, title=sys._getframe().f_code.co_name)

        return

    def updatePoolTeams(self, suppress_prompt: bool=False, batch: bool=False):

        try:

            if suppress_prompt is True:
                response = 1
            else:
                response = self.askToScrapeWebPage()

            if batch:
                logger = logging.getLogger(__name__)
                dialog = None
            else:
                # layout the progress dialog
                layout = [
                    [
                        sg.Text(f'Update Pool Teams for "{self.web_host}"...', size=(60,2), key='-PROG-')
                    ],
                    [
                        sg.Cancel()
                    ],
                ]
                # create the dialog
                dialog = sg.Window(f'Update Pool Teams from "{self.web_host}"', layout, finalize=True, modal=True)

            if response == 1:
                fantrax = Fantrax(pool_id=self.id, league_id=self.league_id, season_id=self.season_id)
                dfPoolTeams = fantrax.scrapePoolTeams(dialog=dialog)
                del fantrax
                # dfPoolTeams.to_sql('dfPoolTeams', con=get_db_connection(), index=False, if_exists='replace')
            elif response == 2:
                msg = 'Getting pool teams data from database...'
                if batch:
                    logger.debug(msg)
                else:
                    dialog['-PROG-'].update(msg)
                    event, values = dialog.read(timeout=2)
                    if event == 'Cancel' or event == sg.WIN_CLOSED:
                        return
                dfPoolTeams = pd.read_sql('select * from dfPoolTeams',con=get_db_connection())
            else:
                return

            if len(dfPoolTeams.index) == 0:
                if batch:
                    logger.debug('No pool teams found. Returning...')
                return

            if response == 1:
                dfPoolTeams.to_sql('dfPoolTeams', con=get_db_connection(), index=False, if_exists='replace')

            msg = f'Rosters collected for {len(dfPoolTeams.index)} pool teams. Writing to database...'
            if batch:
                logger.debug(msg)
            else:
                dialog['-PROG-'].update(msg)
                event, values = dialog.read(timeout=2)
                if event == 'Cancel' or event == sg.WIN_CLOSED:
                    return

            if batch:
                logger.debug(f'Updating pool teams for "{self.web_host}"')
            # Insert/update pool teams
            fantrax = Fantrax(pool_id=self.id, league_id=self.league_id, season_id=self.season_id)
            fantrax.updatePoolTeams(pool=self, df=dfPoolTeams, batch=True)

        except Exception as e:
            msg = f'Error in {sys._getframe().f_code.co_name}: {e}'
            if batch:
                logger.error(rmsg)
            else:
                sg.popup_error(msg)

        finally:
            msg = 'Update of pool teams completed...'
            if batch:
                logger.debug(msg)
            else:
                dialog.close()
                sg.popup_notify(msg, title=sys._getframe().f_code.co_name)

        return

    def updatePoolStandingsGainLoss(self, batch: bool=False):

        try:

            if batch:
                logger = logging.getLogger(__name__)
                dialog = None
            else:
                # layout the progress dialog
                layout = [
                    [
                        sg.Text(f'Update Standings Gain/Loss for "{self.web_host}"...', size=(60,2), key='-PROG-')
                    ],
                    [
                        sg.Cancel()
                    ],
                ]
                # create the dialog
                dialog = sg.Window(f'Update Standings Gain/Loss from "{self.web_host}"', layout, finalize=True, modal=True)
                logger = None

            fantrax = Fantrax(pool_id=self.id, league_id=self.league_id, season_id=self.season_id)
            dfStandingsStats = fantrax.scrapePoolStandingsStats(dialog=dialog)
            del fantrax

            if len(dfStandingsStats.index) == 0:
                if batch:
                    logger.debug('No pool standings stats found. Returning...')
                return

            msg = f'Standings statistics collected. Writing to database...'
            if batch:
                logger.debug(msg)
            else:
                dialog['-PROG-'].update(msg)
                event, values = dialog.read(timeout=2)
                if event == 'Cancel' or event == sg.WIN_CLOSED:
                    return

            # write to database
            dfStandingsStats.to_sql('dfStandingsStats', con=get_db_connection(), index=False, if_exists='replace')

            # write to Excel
            excelSheets = [
                {
                    'df': dfStandingsStats,
                    'excelSheet': 'Stats',
                    'min_row': 2,
                    'max_row': 12,
                    'min_col': 1
                },
            ]
            # self.writeToExcel(df=dfStandingsStats, excelFile='Category Standings Gain & Loss.xlsx', excelSheet='Stats', min_row=2, max_row=12, min_col=1, batch=batch, logger=logger, dialog=dialog)
            self.writeToExcel(excelFile="Category Standings Gain & Loss.xlsx", excelSheets=excelSheets, batch=batch, logger=logger, dialog=dialog)

        except Exception as e:
            msg = f'Error in {sys._getframe().f_code.co_name}: {e}'
            if batch:
                logger.error(msg)
            else:
                sg.popup_error(msg)

        finally:
            msg = 'Update of pool standings statistics completed...'
            if batch:
                logger.debug(msg)
            else:
                dialog.close()
                sg.popup_notify(msg, title=sys._getframe().f_code.co_name)

        return

    def updatePoolTeamServiceTimes(self, batch: bool=False):

        try:

            if batch:
                logger = logging.getLogger(__name__)
                dialog = None
            else:
                # layout the progress dialog
                layout = [
                    [
                        sg.Text(f'Update Pool Team Service Times for "{self.web_host}"...', size=(60,2), key='-PROG-')
                    ],
                    [
                        sg.Cancel()
                    ],
                ]
                # create the dialog
                dialog = sg.Window(f'Update Pool Team Service Times from "{self.web_host}"', layout, finalize=True, modal=True)
                logger = None

            fantrax = Fantrax(pool_id=self.id, league_id=self.league_id, season_id=self.season_id)
            dfPlayerServiceTimes = fantrax.scrapePoolTeamsServiceTime(dialog=dialog)
            del fantrax

            if len(dfPlayerServiceTimes.index) == 0:
                if batch:
                    logger.debug('No pool team service times found. Returning...')
                return

            msg = f'Pool Team Service Times collected. Updating dataframe...'
            if batch:
                logger.debug(msg)
            else:
                dialog['-PROG-'].update(msg)
                event, values = dialog.read(timeout=2)
                if event == 'Cancel' or event == sg.WIN_CLOSED:
                    return

            # add player id column
            dfPlayerServiceTimes['player_id'] = assign_player_ids(df=dfPlayerServiceTimes, player_name='player', nhl_team='team', pos_code='pos')

            #########################################################################################################
            # add pre-draft keeper columns
            sql = f'select player_id, case when keeper = "m" then "MIN" else "Yes" end as pre_draft_keeper, pool_team as pre_draft_manager from KeeperListsArchive where pool_id={self.id}'
            df_keeper_list = pd.read_sql(sql, con=get_db_connection())

            # add pre-draft keeper columns to df
            dfPlayerServiceTimes.set_index(['player_id'], inplace=True)
            df_keeper_list.set_index(['player_id'], inplace=True)

            dfPlayerServiceTimes['keeper'] = dfPlayerServiceTimes.apply(lambda row: 1 if row.name in df_keeper_list.index else 0, axis=1)

            dfPlayerServiceTimes['orig_keeper'] = dfPlayerServiceTimes.apply(lambda row: 1 if row.name in df_keeper_list.index and row['manager'] == df_keeper_list.loc[row.name, 'pre_draft_manager'] else 0, axis=1)

            dfPlayerServiceTimes['other_keeper'] = dfPlayerServiceTimes.apply(lambda row: 1 if row.name in df_keeper_list.index and row['manager'] != df_keeper_list.loc[row.name, 'pre_draft_manager'] else 0, axis=1)

            dfPlayerServiceTimes.reset_index(inplace=True)
            #########################################################################################################

            # list of period columns (i.e., 1, 2, 3, ...)
            period_columns = [col for col in dfPlayerServiceTimes.columns if col.isdigit()]

            # on-original-roster column
            dfPlayerServiceTimes['original'] = dfPlayerServiceTimes[period_columns].iloc[:, 0].apply(lambda x: 1 if x != '--' else 0)

            # on-current-roster column
            dfPlayerServiceTimes['current'] = dfPlayerServiceTimes[period_columns].iloc[:, -1].apply(lambda x: 1 if x != '--' else 0)

            # static column (players who were on the original roster and on the current roster)
            dfPlayerServiceTimes['static'] = dfPlayerServiceTimes.apply(lambda row: 1 if row[period_columns[0]] != '--' and row[period_columns[-1]] != '--' else 0, axis=1)

            # added column (players who were not on the original roster but now on the current roster)
            dfPlayerServiceTimes['added'] = dfPlayerServiceTimes.apply(lambda row: 1 if row[period_columns[0]] == '--' and row[period_columns[-1]] != '--' else 0, axis=1)

            # moved column (players who were on the original roster but not on the current roster)
            dfPlayerServiceTimes['moved'] = dfPlayerServiceTimes.apply(lambda row: 1 if row[period_columns[0]] != '--' and row[period_columns[-1]] == '--' else 0, axis=1)

            # on-fill-in-roster column
            dfPlayerServiceTimes['fill_in'] = dfPlayerServiceTimes.apply(lambda row: 1 if row[period_columns[0]] == '--' and row[period_columns[-1]] == '--' else 0, axis=1)

            # player as Skt column
            dfPlayerServiceTimes['as_skt'] = dfPlayerServiceTimes.apply(lambda row: 1 if any(row[col] == 'Skt' for col in period_columns) else 0, axis=1)

            # player as 'Min' column
            dfPlayerServiceTimes['as_min'] = dfPlayerServiceTimes.apply(lambda row: 1 if any(row[col] == 'Min' for col in period_columns) else 0, axis=1)

            # player always 'Min' column
            dfPlayerServiceTimes['always_min'] = dfPlayerServiceTimes.apply(lambda row: 1 if any(row[col] == 'Min' for col in period_columns) and all(row[col] in ['Min', '--'] for col in period_columns) else 0, axis=1)

            # player as 'Min' now column
            dfPlayerServiceTimes['min_now'] = dfPlayerServiceTimes.apply(lambda row: 1 if row[period_columns[-1]] == 'Min' else 0, axis=1)

            # Reorder columns to position 'player_id' and 'current' after 'team'
            cols = dfPlayerServiceTimes.columns.tolist()
            player_id_index = cols.index('player_id')
            cols.insert(player_id_index + 1, cols.pop(cols.index('player_id')))
            team_abbr_index = cols.index('team')
            # and in reverse order to insert 'original' as the first column
            cols.insert(team_abbr_index + 1, cols.pop(cols.index('min_now')))
            cols.insert(team_abbr_index + 1, cols.pop(cols.index('always_min')))
            cols.insert(team_abbr_index + 1, cols.pop(cols.index('as_min')))
            cols.insert(team_abbr_index + 1, cols.pop(cols.index('as_skt')))
            cols.insert(team_abbr_index + 1, cols.pop(cols.index('fill_in')))
            cols.insert(team_abbr_index + 1, cols.pop(cols.index('added')))
            cols.insert(team_abbr_index + 1, cols.pop(cols.index('moved')))
            cols.insert(team_abbr_index + 1, cols.pop(cols.index('static')))
            cols.insert(team_abbr_index + 1, cols.pop(cols.index('original')))
            cols.insert(team_abbr_index + 1, cols.pop(cols.index('current')))
            cols.insert(team_abbr_index + 1, cols.pop(cols.index('other_keeper')))
            cols.insert(team_abbr_index + 1, cols.pop(cols.index('orig_keeper')))
            cols.insert(team_abbr_index + 1, cols.pop(cols.index('keeper')))
            dfPlayerServiceTimes = dfPlayerServiceTimes[cols]

            msg = f' Writing Pool Team Service Times to database...'
            if batch:
                logger.debug(msg)
            else:
                dialog['-PROG-'].update(msg)
                event, values = dialog.read(timeout=2)
                if event == 'Cancel' or event == sg.WIN_CLOSED:
                    return

            # write to database
            dfPlayerServiceTimes.to_sql('dfPlayerServiceTimes', con=get_db_connection(), index=False, if_exists='replace')

            # # write to Excel
            # self.writeToExcel(df=dfPlayerServiceTimes, excelFile="Manager Player Service Times.xlsx", excelSheet='Service Time Details', min_row=2, max_row=None, min_col=1, batch=batch, logger=logger, dialog=dialog)

            ########################################################################

            msg = f'Creating Pool Team Service Times Summary dataframe...'
            if batch:
                logger.debug(msg)
            else:
                dialog['-PROG-'].update(msg)
                event, values = dialog.read(timeout=2)
                if event == 'Cancel' or event == sg.WIN_CLOSED:
                    return

            # add column for full service time
            dfPlayerServiceTimes['full'] = dfPlayerServiceTimes[period_columns].apply(lambda row: 1 if all(val != '--' for val in row) else 0, axis=1)

            # Create a summary dataframe
            dfPlayerServiceTimesSummary = dfPlayerServiceTimes.groupby('manager').agg(
                all=('player_id', 'count'),
                all_Fs=('pos', lambda x: (x == 'F').sum()),
                all_Ds=('pos', lambda x: (x == 'D').sum()),
                all_Gs=('pos', lambda x: (x == 'G').sum()),
                current=('current', lambda x: (x == 1).sum()),
                curr_Fs=('current', lambda x: ((dfPlayerServiceTimes['pos'] == 'F') & (x == 1)).sum()),
                curr_Ds=('current', lambda x: ((dfPlayerServiceTimes['pos'] == 'D') & (x == 1)).sum()),
                curr_Gs=('current', lambda x: ((dfPlayerServiceTimes['pos'] == 'G') & (x == 1)).sum()),
                keeper=('keeper', lambda x: ((dfPlayerServiceTimes['current'] == 1) & (x == 1)).sum()),
                keeper_Fs=('keeper', lambda x: ((dfPlayerServiceTimes['current'] == 1) & (dfPlayerServiceTimes['pos'] == 'F') & (x == 1)).sum()),
                keeper_Ds=('keeper', lambda x: ((dfPlayerServiceTimes['current'] == 1) & (dfPlayerServiceTimes['pos'] == 'D') & (x == 1)).sum()),
                keeper_Gs=('keeper', lambda x: ((dfPlayerServiceTimes['current'] == 1) & (dfPlayerServiceTimes['pos'] == 'G') & (x == 1)).sum()),
                # orig_keeper=('orig_keeper', lambda x: ((dfPlayerServiceTimes['current'] == 1) & (x == 1)).sum()),
                # orig_keeper_Fs=('orig_keeper', lambda x: ((dfPlayerServiceTimes['current'] == 1) & (dfPlayerServiceTimes['pos'] == 'F') & (x == 1)).sum()),
                # orig_keeper_Ds=('orig_keeper', lambda x: ((dfPlayerServiceTimes['current'] == 1) & (dfPlayerServiceTimes['pos'] == 'D') & (x == 1)).sum()),
                # orig_keeper_Gs=('orig_keeper', lambda x: ((dfPlayerServiceTimes['current'] == 1) & (dfPlayerServiceTimes['pos'] == 'G') & (x == 1)).sum()),
                # other_keeper=('other_keeper', lambda x: ((dfPlayerServiceTimes['current'] == 1) & (x == 1)).sum()),
                # other_keeper_Fs=('other_keeper', lambda x: ((dfPlayerServiceTimes['current'] == 1) & (dfPlayerServiceTimes['pos'] == 'F') & (x == 1)).sum()),
                # other_keeper_Ds=('other_keeper', lambda x: ((dfPlayerServiceTimes['current'] == 1) & (dfPlayerServiceTimes['pos'] == 'D') & (x == 1)).sum()),
                # other_keeper_Gs=('other_keeper', lambda x: ((dfPlayerServiceTimes['current'] == 1) & (dfPlayerServiceTimes['pos'] == 'G') & (x == 1)).sum()),
                original=('original', lambda x: (x == 1).sum()),
                orig_Fs=('original', lambda x: ((dfPlayerServiceTimes['pos'] == 'F') & (x == 1)).sum()),
                orig_Ds=('original', lambda x: ((dfPlayerServiceTimes['pos'] == 'D') & (x == 1)).sum()),
                orig_Gs=('original', lambda x: ((dfPlayerServiceTimes['pos'] == 'G') & (x == 1)).sum()),
                static=('static', lambda x: (x == 1).sum()),
                static_Fs=('static', lambda x: ((dfPlayerServiceTimes['pos'] == 'F') & (x == 1)).sum()),
                static_Ds=('static', lambda x: ((dfPlayerServiceTimes['pos'] == 'D') & (x == 1)).sum()),
                static_Gs=('static', lambda x: ((dfPlayerServiceTimes['pos'] == 'G') & (x == 1)).sum()),
                added=('added', lambda x: (x == 1).sum()),
                added_Fs=('added', lambda x: ((dfPlayerServiceTimes['pos'] == 'F') & (x == 1)).sum()),
                added_Ds=('added', lambda x: ((dfPlayerServiceTimes['pos'] == 'D') & (x == 1)).sum()),
                added_Gs=('added', lambda x: ((dfPlayerServiceTimes['pos'] == 'G') & (x == 1)).sum()),
                moved=('moved', lambda x: (x == 1).sum()),
                moved_Fs=('moved', lambda x: ((dfPlayerServiceTimes['pos'] == 'F') & (x == 1)).sum()),
                moved_Ds=('moved', lambda x: ((dfPlayerServiceTimes['pos'] == 'D') & (x == 1)).sum()),
                moved_Gs=('moved', lambda x: ((dfPlayerServiceTimes['pos'] == 'G') & (x == 1)).sum()),
                fill_in=('fill_in', lambda x: (x == 1).sum()),
                fill_in_Fs=('fill_in', lambda x: ((dfPlayerServiceTimes['pos'] == 'F') & (x == 1)).sum()),
                fill_in_Ds=('fill_in', lambda x: ((dfPlayerServiceTimes['pos'] == 'D') & (x == 1)).sum()),
                fill_in_Gs=('fill_in', lambda x: ((dfPlayerServiceTimes['pos'] == 'G') & (x == 1)).sum()),
                as_skt=('as_skt', lambda x: (x == 1).sum()),
                f_as_skt=('as_skt', lambda x: ((dfPlayerServiceTimes['pos'] == 'F') & (x == 1)).sum()),
                d_as_skt=('as_skt', lambda x: ((dfPlayerServiceTimes['pos'] == 'D') & (x == 1)).sum()),
                as_min=('as_min', lambda x: (x == 1).sum()),
                f_as_min=('as_min', lambda x: ((dfPlayerServiceTimes['pos'] == 'F') & (x == 1)).sum()),
                d_as_min=('as_min', lambda x: ((dfPlayerServiceTimes['pos'] == 'D') & (x == 1)).sum()),
                g_as_min=('as_min', lambda x: ((dfPlayerServiceTimes['pos'] == 'G') & (x == 1)).sum()),
                always_min=('always_min', lambda x: (x == 1).sum()),
                f_always_min=('always_min', lambda x: ((dfPlayerServiceTimes['pos'] == 'F') & (x == 1)).sum()),
                d_always_min=('always_min', lambda x: ((dfPlayerServiceTimes['pos'] == 'D') & (x == 1)).sum()),
                g_always_min=('always_min', lambda x: ((dfPlayerServiceTimes['pos'] == 'G') & (x == 1)).sum()),
                min_now=('min_now', lambda x: (x == 1).sum()),
                f_min_now=('min_now', lambda x: ((dfPlayerServiceTimes['pos'] == 'F') & (x == 1)).sum()),
                d_min_now=('min_now', lambda x: ((dfPlayerServiceTimes['pos'] == 'D') & (x == 1)).sum()),
                g_min_now=('min_now', lambda x: ((dfPlayerServiceTimes['pos'] == 'G') & (x == 1)).sum()),
           ).reset_index()

            msg = f'Writing Pool Team Service Times Summary to database...'
            if batch:
                logger.debug(msg)
            else:
                dialog['-PROG-'].update(msg)
                event, values = dialog.read(timeout=2)
                if event == 'Cancel' or event == sg.WIN_CLOSED:
                    return

            # Add the summary dataframe to the database
            dfPlayerServiceTimesSummary.to_sql('dfPlayerServiceTimesSummary', con=get_db_connection(), index=False, if_exists='replace')

            # write to Excel
            excelSheets = [
                {
                    'df': dfPlayerServiceTimes,
                    'excelSheet': 'Service Time Details',
                    'min_row': 2,
                    'max_row': None,
                    'min_col': 1
                },
                {
                    'df': dfPlayerServiceTimesSummary,
                    'excelSheet': 'Service Time Summary',
                    'min_row': 3,
                    'max_row': None,
                    'min_col': 1
                }
            ]
            # self.writeToExcel(df=dfPlayerServiceTimes, excelFile="Manager Player Service Times.xlsx", excelSheet='Service Time Details', min_row=2, max_row=None, min_col=1, batch=batch, logger=logger, dialog=dialog)
            # self.writeToExcel(df=dfPlayerServiceTimesSummary, excelFile="Manager Player Service Times.xlsx", excelSheet='Service Time Summary', min_row=3, max_row=None, min_col=1, batch=batch, logger=logger, dialog=dialog)
            self.writeToExcel(excelFile="Manager Player Service Times.xlsx", excelSheets=excelSheets, batch=batch, logger=logger, dialog=dialog)

        except Exception as e:
            msg = f'Error in {sys._getframe().f_code.co_name}: {e}'
            if batch:
                logger.error(msg)
            else:
                sg.popup_error(msg)

        finally:
            msg = 'Update of pool team service times completed...'
            if batch:
                logger.debug(msg)
            else:
                dialog.close()
                sg.popup_notify(msg, title=sys._getframe().f_code.co_name)

        return

    def window(self):

        global season
        global window

        try:

            # get stats for all players
            self.getPlayerStats()

            window = sg.Window('NHL Pool', self.layout_window(), font=("Helvetica", 12), resizable=True, return_keyboard_events=True, finalize=True, size=(1000,670))

            ft_pt_mclb = window['__FT_PT_MCLB__']
            ft_ptr_mclb = window['__FT_PTR_MCLB__']
            ft_pt_mclb_container = window['__FT_PT_MCLB_CNTNR__']
            ft_ptr_mclb_container = window['__FT_PTR_MCLB_CNTNR__']

            # Need to capture single-click of left-mouse-button
            def lmb_single_click(event):
                window['LEFT_MOUSE_BUTTON']._ClickHandler('<ButtonRelease-1>')
            ft_pt_mclb.Widget.bind('<ButtonRelease-1>', lmb_single_click)

            # Need to capture single-click of right-mouse-button
            ft_ptr_mclb.bind('<ButtonPress-3>', '+RIGHT_CLICK+')

            pool_teams_mclb_selected_row = 0
            pool_team_roster_mclb_selected_row = 0

            update_pool_teams_tab = True

            refresh_pool_team_config = True
            refresh_pool_team_roster_config = True

            config_sort_column_name = None

            while True:

                try:

                    pt_mclb = ft_pt_mclb
                    ptr_mclb = ft_ptr_mclb

                    ###############################################################################
                    # Pools teams tab
                    ###############################################################################
                    if update_pool_teams_tab is True:

                        if refresh_pool_team_config is True:

                            # set mclb configs
                            self.config_pool_team()
                            self.config_pool_team_roster()

                            # hide all mclbs, to start
                            ft_pt_mclb_container.update(visible=False)
                            ft_ptr_mclb_container.update(visible=False)

                            # Update Pool Teams tab
                            # if season.SEASON_HAS_STARTED is True and season.type == 'P':
                            #     next_season_id = season.getNextSeasonID()
                            #     next_season_pool = HockeyPool().fetch(**{'Criteria': [['season_id', '==', next_season_id]]})
                            #     df = self.get_pool_teams_as_df(pool=next_season_pool)
                            # else:
                            #     df = self.get_pool_teams_as_df(pool=self)
                            df = self.get_pool_teams_as_df(pool=self)

                            if config_sort_column_name is None:
                                df.sort_values(by=['points', 'name'], ascending=[False, True], inplace=True)
                            else:
                                if config_sort_column_name == 'name':
                                    df.sort_values(by='name', inplace=True)
                                else:
                                    df.sort_values(by=[config_sort_column_name, 'name'], ascending=[False, True], inplace=True)

                            config_sort_column_name = None

                            pool_teams_list = self.getMCLBData(config=pool_team_config, df=df)
                            # hide not-wanted mclb
                            ft_pt_mclb_container.update(visible=True)
                            # update mclb data
                            pt_mclb.update(values=[[pool_team[i] for i,x in enumerate(pool_team)] for pool_team in pool_teams_list])

                            # format table columns
                            self.table_container_formatting(config_columns=pool_team_config['columns'], mclb=pt_mclb)
                            refresh_pool_team_config = False

                        # Update Pool Team players
                        # get index for pool team's "id" column
                        idx = [i for i, x in enumerate(pool_team_config['columns']) if 'table column' in x and x['table column']=='id'][0]
                        # get pool team name
                        poolteams = pt_mclb.get()
                        if len(poolteams) > 0:
                            poolteam_id = pt_mclb.get()[pool_teams_mclb_selected_row][idx]
                        else:
                            poolteam_id = 0
                        # get roster players & statistics
                        kwargs = {'Criteria': [['poolteam_id', '==', poolteam_id]]}
                        players: List[int] = PoolTeamRoster().fetch_many(**kwargs)
                        player_ids = [x.player_id for x in players]

                        df = df_player_stats.query(f'player_id in @player_ids').copy()

                        if config_sort_column_name is None:
                            df.sort_values(by=['points', 'games'], inplace=True)
                        else:
                            if config_sort_column_name == 'name':
                                df.sort_values(by='name', inplace=True)
                            else:
                                sequence = [False, True]
                                df.sort_values(by=[config_sort_column_name, 'name'], ascending=sequence, inplace=True)

                        player_stats = self.getMCLBData(config=pool_team_roster_config, df=df, sort_override=(config_sort_column_name != None))
                        # player_stats = self.getMCLBData(config=pool_team_roster_config, df=df)

                        # hide not-wanted mclb
                        ft_ptr_mclb_container.update(visible=True)
                        # update mclb data
                        ptr_mclb.update(values=[[stat[i] for i,x in enumerate(stat)] for stat in player_stats])

                        if refresh_pool_team_roster_config is True:
                            # format table columns
                            self.table_container_formatting(config_columns=pool_team_roster_config['columns'], mclb=ptr_mclb)
                            refresh_pool_team_roster_config = False

                        update_pool_teams_tab = False

                        # Set pool teams mclb focus, row focus & selected row
                        pt_mclb.set_focus()
                        if len(pt_mclb.get()) > 0:
                            pt_mclb.update(select_rows=[pool_teams_mclb_selected_row])
                            pt_mclb.SelectedRows = [pool_teams_mclb_selected_row]
                            # pt_mclb.Widget.see(pool_teams_mclb_selected_row - 1)
                            pt_mclb.Widget.focus(pool_teams_mclb_selected_row + 1)

                        if len(ptr_mclb.get()) > 0:
                            ptr_mclb.update(select_rows=[pool_team_roster_mclb_selected_row])
                            ptr_mclb.SelectedRows = [pool_team_roster_mclb_selected_row]
                            # ptr_mclb.Widget.see(pool_team_roster_mclb_selected_row - 1)
                            ptr_mclb.Widget.focus(pool_team_roster_mclb_selected_row + 1)

                    ###############################################################################
                    # Read window
                    ###############################################################################

                    # save mclb row selections to compare when left-mouse-button, or up
                    # down arrows used to select rows on master-detail tabs
                    current_pt_mclb_selection = pt_mclb.SelectedRows
                    current_ptr_mclb_selection = ptr_mclb.SelectedRows

                    event, values = window.Read()

                    #  File Menu
                    if event in (None, 'Exit'):
                        break

                    elif event in ['MouseWheel:Down', 'MouseWheel:Up']:
                        continue

                    elif event == '__SEASON_COMBO__':

                        # set season and return new hockey pool
                        self = self.set_season(event_values=values)

                        # re-get player stats for new pool
                        self.getPlayerStats()

                        pool_teams_mclb_selected_row = 0

                        update_pool_teams_tab = True

                        refresh_pool_team_config = True
                        refresh_pool_team_roster_config = True

                    elif event == '__POOL_COMBO__':

                        self.set_pool(event_values=values)

                        pool_teams_mclb_selected_row = 0

                        config_sort_column_name = None

                        update_pool_teams_tab = True

                        refresh_pool_team_config = True
                        refresh_pool_team_roster_config = True

                    elif '__FT_PT_MCLB__' in event:
                        selections = []
                        if type(event) == tuple:
                            (selected_row, selected_column) = event[2]
                            if selected_row == -1:
                                idx = [i for i, x in enumerate(pool_team_config['columns']) if 'visible' not in x or x['visible']==True]
                                column: Dict = pool_team_config['columns'][idx[selected_column]]
                                config_sort_column_name = column['table column'] if 'table column' in column else column['runtime column']
                                update_pool_teams_tab = True
                                refresh_pool_team_config = True
                                refresh_pool_team_roster_config = True
                            pool_team_roster_mclb_selected_row = 0
                            continue
                        else:
                            selections = [selected_row]

                        for selection in selections:
                            sel_pool_team = pt_mclb.get()[selection]
                            idx = [i for i, x in enumerate(pool_team_config['columns']) if 'table column' in x and x['table column']=='id'][0]
                            kwargs = {'Criteria': [['id', '==', sel_pool_team[idx]]]}
                            poolTeam = PoolTeam().fetch(**kwargs)
                            poolTeam.dialog()

                            self.config_pool_team()
                            self.config_pool_team_roster()

                            pool_teams_mclb_selected_row = selection

                        update_pool_teams_tab = True
                        refresh_pool_team_config = True
                        refresh_pool_team_roster_config = True

                    # # elif (type(event) == tuple and event[0] == '__FT_PTR_MCLB__') or event == '__FT_PTR_MCLB__':
                    # elif type(event) == tuple and event[0] == '__FT_PTR_MCLB__':
                    #     selections = []
                    #     mclb = ft_ptr_mclb
                    #     if type(event) == tuple:
                    #         (selected_row, selected_column) = event[2]
                    #         if selected_row == -1:
                    #             idx = [i for i, x in enumerate(pool_team_roster_config['columns']) if 'visible' not in x or x['visible']==True]
                    #             column: Dict = pool_team_roster_config['columns'][idx[selected_column]]
                    #             config_sort_column_name = column['table column'] if 'table column' in column else column['runtime column']
                    #             update_pool_teams_tab = True
                    #             # refresh_pool_team_config = True
                    #             refresh_pool_team_roster_config = True
                    #             continue
                    #         else:
                    #             selections = [selected_row]

                    #     for selection in selections:
                    #         sel_player = mclb.get()[selection]
                    #         idx = [i for i, x in enumerate(pool_team_roster_config['columns']) if 'table column' in x and x['table column']=='player_id'][0]
                    #         kwargs = {f'id': sel_player[idx]}
                    #         player = Player().fetch(**kwargs)
                    #         player.window(self.season_id)

                    elif event == 'Player Bio':
                        for selection in ft_ptr_mclb.SelectedRows:
                            sel_player = ft_ptr_mclb.get()[selection]
                            idx = [i for i, x in enumerate(pool_team_roster_config['columns']) if 'table column' in x and x['table column']=='player_id'][0]
                            kwargs = {f'id': sel_player[idx]}
                            player = Player().fetch(**kwargs)
                            player.window(self.season_id)

                    elif event == 'Update Player':
                        for selection in mclb.SelectedRows:
                            sel_poolteam = pt_mclb.get()[pt_mclb.SelectedRows[0]]
                            pt_idx = [i for i, x in enumerate(pool_team_config['columns']) if 'table column' in x and x['table column']=='id'][0]
                            sel_player = mclb.get()[selection]
                            p_idx = [i for i, x in enumerate(pool_team_roster_config['columns']) if 'table column' in x and x['table column']=='player_id'][0]
                            kwargs = {'poolteam_id': sel_poolteam[pt_idx], 'player_id': sel_player[p_idx]}
                            roster_player = PoolTeamRoster().fetch(**kwargs)
                            roster_player.dialog()

                    elif event in ('Mark as Keeper', 'Mark as Minor', 'Unmark as Keeper/Minor'):
                        if self.web_host == 'Fantrax' and values['Tab'] == '__POOL_TEAMS_TAB__':
                            mclb = window['__FT_PTR_MCLB__']
                        else:
                            continue

                        for selection in mclb.SelectedRows:
                            sel_poolteam = pt_mclb.get()[pt_mclb.SelectedRows[0]]
                            pt_idx = [i for i, x in enumerate(pool_team_config['columns']) if 'table column' in x and x['table column']=='id'][0]
                            sel_player = mclb.get()[selection]
                            p_idx = [i for i, x in enumerate(pool_team_roster_config['columns']) if 'table column' in x and x['table column']=='player_id'][0]
                            kwargs = {'poolteam_id': sel_poolteam[pt_idx], 'player_id': sel_player[p_idx]}
                            roster_player = PoolTeamRoster().fetch(**kwargs)
                            if roster_player.keeper in ('y', 'm') and event == 'Unmark as Keeper/Minor':
                                roster_player.keeper = ''
                            elif roster_player.keeper in ('', 'm', None) and event == 'Mark as Keeper':
                                roster_player.keeper = 'y'
                            elif roster_player.keeper in ('','y', None) and event == 'Mark as Minor':
                                roster_player.keeper = 'm'
                            roster_player.persist()

                            pool_team_roster_mclb_selected_row = selection

                        update_pool_teams_tab = True
                        # re-get player stats for changed "keeper" flags
                        self.getPlayerStats()
                        refresh_pool_team_roster_config = True

                    elif event == 'Refresh':
                        # get stats for all players
                        self.getPlayerStats()
                        update_pool_teams_tab = True
                        refresh_pool_team_config = True
                        refresh_pool_team_roster_config = True

                    # NOTE: Need the ending "," to stop "TypeError: 'in <string>' requires string as left operand, not tuple" message
                    elif event == 'Refresh Roster':
                        if self.web_host == 'Fantrax' and values['Tab'] == '__POOL_TEAMS_TAB__':
                            mclb = window['__FT_PT_MCLB__']
                        else:
                            continue

                        for selection in mclb.SelectedRows:
                            sel_poolteam = pt_mclb.get()[selection]
                            pt_idx = [i for i, x in enumerate(pool_team_config['columns']) if 'table column' in x and x['table column']=='name'][0]
                            # kwargs = {'Criteria': [['id', '==', sel_poolteam[pt_idx]]]}
                            # pool_team = PoolTeam().fetch(**kwargs)
                            self.updatePoolTeamRosters(suppress_prompt=True, pool_teams=[sel_poolteam[pt_idx]])

                            # pool_team_roster_mclb_selected_row = selection

                        update_pool_teams_tab = True
                        self.getPlayerStats()
                        refresh_pool_team_roster_config = True

                    elif event == 'Get Period Rosters':
                        if self.web_host == 'Fantrax' and values['Tab'] == '__POOL_TEAMS_TAB__':
                            mclb = window['__FT_PT_MCLB__']
                        else:
                            continue

                        for selection in mclb.SelectedRows:
                            sel_poolteam = pt_mclb.get()[selection]
                            pt_idx = [i for i, x in enumerate(pool_team_config['columns']) if 'table column' in x and x['table column']=='name'][0]
                            # kwargs = {'Criteria': [['id', '==', sel_poolteam[pt_idx]]]}
                            # pool_team = PoolTeam().fetch(**kwargs)
                            self.getPoolTeamRostersByPeriod(pool_teams=[sel_poolteam[pt_idx]])

                    elif event == 'Remove Player':
                        if self.web_host == 'Fantrax' and values['Tab'] == '__POOL_TEAMS_TAB__':
                            mclb = window['__FT_PTR_MCLB__']
                        else:
                            continue

                        for selection in mclb.SelectedRows:
                            sel_poolteam = pt_mclb.get()[pt_mclb.SelectedRows[0]]
                            pt_idx = [i for i, x in enumerate(pool_team_config['columns']) if 'table column' in x and x['table column']=='id'][0]
                            sel_player = mclb.get()[selection]
                            p_idx = [i for i, x in enumerate(pool_team_roster_config['columns']) if 'table column' in x and x['table column']=='player_id'][0]
                            kwargs = {'poolteam_id': sel_poolteam[pt_idx], 'player_id': sel_player[p_idx]}
                            roster_player = PoolTeamRoster().fetch(**kwargs)
                            roster_player.destroy(**kwargs)
                            update_pool_teams_tab = True
                            refresh_pool_team_roster_config = True

                    elif event == 'Delete:46':
                        continue

                    elif event == 'Export to Excel':
                        self.export_to_excel(mclb=pps_mclb)

                    elif event == '__NEW_POOL_TEAM__':
                        poolTeam = PoolTeam()
                        poolTeam.pool_id = self.id
                        poolTeam.dialog()
                        selections = values['__FT_PT_MCLB__']
                        pool_teams_mclb_selected_row = len(values['__FT_PT_MCLB__']) - 1
                        update_pool_teams_tab = True
                        refresh_pool_team_config = True

                    elif event == '__NEW_POOL_TEAM_PLAYER__':
                        poolteam_player = PoolTeamRoster()

                        poolteam_player.poolteam_id = poolteam_id
                        poolteam_player.dialog()
                        update_pool_teams_tab = True
                        refresh_pool_team_config = True
                        refresh_pool_team_roster_config = True

                    elif event == 'Season_ID':
                        self.season_id = values['Season_ID']
                        season = Season().getSeason(id=self.season_id)
                        window['Season_Name'].update(value=season.name)
                        window['Season_Start'].update(value=season.start_date)
                        window['Season_End'].update(value=season.end_date)
                        window['Number_Of_Games'].update(value=season.number_of_games)
                        window['Weeks_In_Season'].update(value=season.weeks)
                        window['count_of_total_game_dates'].update(value=season.count_of_total_game_dates)
                        window['count_of_completed_game_dates'].update(value=season.count_of_completed_game_dates)
                        self.getPlayerStats()
                        update_pool_teams_tab = True

                    elif event in ['LEFT_MOUSE_BUTTON', 'Down:40', 'Up:38']:
                        # pp_mclb.TKTreeview.column(2)
                        # function variables
                        # 'width':55
                        # 'minwidth':10
                        # 'stretch':0
                        # 'anchor':'center'
                        # 'id':'Team'
                        # len():5
                        # pp_mclb.TKTreeview.winfo_geometry()
                        # '1443x472+0+0'
                        if values['Tab'] == '__POOL_TEAMS_TAB__':
                            selections = []
                            if current_pt_mclb_selection != values['__FT_PT_MCLB__']:
                                selections = values['__FT_PT_MCLB__']
                                for selection in selections:
                                    pool_teams_mclb_selected_row = selection
                                update_pool_teams_tab = True
                                refresh_pool_team_roster_config = True

                    elif event == '__FT_PTR_MCLB__+RIGHT_CLICK+':
                        def update_right_click_menu(menu, roster_player):
                            if roster_player.keeper == 'y':
                                menu[1] = ['Mark as Minor', 'Unmark as Keeper/Minor', '-', 'Player Bio', '-','Remove Player']
                            elif roster_player.keeper == 'm':
                                menu[1] = ['Mark as Keeper', 'Unmark as Keeper/Minor', '-', 'Player Bio', '-','Remove Player']
                            else:
                                menu[1] = ['Mark as Keeper', 'Mark as Minor', '-', 'Player Bio', '-','Remove Player']

                        right_click_menu = window['__FT_PTR_MCLB__'].RightClickMenu
                        if len(window['__FT_PTR_MCLB__'].SelectedRows) == 1:
                            selection = window['__FT_PTR_MCLB__'].SelectedRows[0]
                            sel_poolteam = window['__FT_PT_MCLB__'].get()[window['__FT_PT_MCLB__'].SelectedRows[0]]
                            pt_idx = [i for i, x in enumerate(pool_team_config['columns']) if 'table column' in x and x['table column']=='id'][0]
                            sel_player = window['__FT_PTR_MCLB__'].get()[selection]
                            p_idx = [i for i, x in enumerate(pool_team_roster_config['columns']) if 'table column' in x and x['table column']=='player_id'][0]
                            kwargs = {'poolteam_id': sel_poolteam[pt_idx], 'player_id': sel_player[p_idx]}
                            roster_player = PoolTeamRoster().fetch(**kwargs)
                            update_right_click_menu(right_click_menu, roster_player)
                        else:
                            right_click_menu[1] = ['Mark as Keeper', 'Mark as Minor', 'Unmark as Keeper/Minor']
                        window['__FT_PTR_MCLB__'].RightClickMenu = right_click_menu

                        ft_ptr_mclb.set_right_click_menu(right_click_menu)

                    elif event == 'Edit_Hockey_Pool_Button':
                        window['Season'].update(disabled=False)
                        window['Season_Start'].update(disabled=False)
                        window['Season_End'].update(disabled=False)
                        window['Weeks_In_Season'].update(disabled=False)
                        window['Edit_Hockey_Pool_Button'].update(visible=False)
                        window['Save_Hockey_Pool_Button'].update(visible=True)
                        window['Cancel_Hockey_Pool_Button'].update(visible=True)

                    elif event == 'Save_Hockey_Pool_Button':
                        self.persist()

                        window['Season'].update(disabled=False)
                        window['Season_Start'].update(disabled=True)
                        window['Season_End'].update(disabled=True)
                        window['Weeks_In_Season'].update(disabled=True)
                        window['Edit_Hockey_Pool_Button'].update(visible=True)
                        window['Save_Hockey_Pool_Button'].update(visible=False)
                        window['Cancel_Hockey_Pool_Button'].update(visible=False)

                    elif event == 'Cancel_Hockey_Pool_Button':
                        window['Season'].update(disabled=False)
                        window['Season_Start'].update(disabled=True)
                        window['Season_End'].update(disabled=True)
                        window['Weeks_In_Season'].update(disabled=True)
                        window['Edit_Hockey_Pool_Button'].update(visible=True)
                        window['Save_Hockey_Pool_Button'].update(visible=False)
                        window['Cancel_Hockey_Pool_Button'].update(visible=False)

                    elif event == 'Get Season Info':
                        self.apiGetSeason(season=season)
                    elif event == 'Get Teams':
                        self.apiGetTeams()
                    elif event == 'Get Players':
                        self.getPlayers()
                    elif event == 'Get Player Stats':
                        self.apiGetPlayerSeasonStats()
                        # get stats for all players
                        self.getPlayerStats()
                        update_pool_teams_tab = True
                        refresh_pool_team_config = True
                        refresh_pool_team_roster_config = True

                    elif event == 'Get MoneyPuck Data':
                        self.getMoneyPuckData(season=season)

                    elif event == 'Update Player Injuries':
                        self.updatePlayerInjuries(suppress_prompt=True)
                        update_pool_teams_tab = True
                        refresh_pool_team_roster_config = True
                        refresh_nhl_team_player_stats_config = True

                    elif event == 'Email NHL Team Transactions':
                        self.email_nhl_team_transactions()

                    elif event == 'Email Starting Goalie Projections':
                        self.email_starting_goalie_projections(pool_id=self.id)

                    elif event == 'Update Pool Teams':
                        self.updatePoolTeams(suppress_prompt=True)
                        pool_teams_mclb_selected_row = 0
                        update_pool_teams_tab = True
                        refresh_pool_team_config = True

                    elif event == 'Update Pool Team Rosters':
                        self.updatePoolTeamRosters(suppress_prompt=True)
                        self.getPlayerStats()
                        pool_teams_mclb_selected_row = 0
                        update_pool_teams_tab = True
                        refresh_pool_team_roster_config = True

                    elif event == 'Update Standings Gain/Loss':
                        self.updatePoolStandingsGainLoss()

                    elif event == 'Update Pool Team Service Times':
                        self.updatePoolTeamServiceTimes()

                    elif event == 'Update Full Team Player Scoring':
                        self.updateFullTeamPlayerScoring()

                    elif event == 'Get Pool Team Rosters, by Period':
                        self.getPoolTeamRostersByPeriod()

                    elif event == 'Update Fantrax Player Info':
                        self.updateFantraxPlayerInfo()

                    elif event == 'Import Watch List':
                        # self.importFantraxWatchList()
                        self.updateFantraxPlayerInfo(watchlist=True)

                    elif event == 'Update Player Lines':
                        self.update_player_lines(season=season)

                    elif event == 'Import Draft Picks':
                        self.import_draft_picks(season=season)

                    elif event == 'Import Transaction History':
                        self.import_transaction_history(season=season)

                    elif event == 'Summarize Trades History':
                        self.summarize_trades_history()

                    elif event == 'Summarize Claims History':
                        self.summarize_claims_history()

                    elif event == 'Archive Keeper Lists':
                        self.archive_keeper_lists(season=season, pool=self)

                    elif event == 'Athletic Import':
                        self.import_athletic_projected_stats(season=season)

                    elif event == 'Dobber Import':
                        self.import_dobber_projected_stats(season=season)

                    elif event == 'Fantrax Import':
                        self.import_fantrax_projected_stats(season=season)

                    elif event == 'Manager Game Pace':
                        ps.manager_game_pace(season=season, pool=self)

                    elif event == 'Position Statistics':
                        self.show_stat_summary_tables(season=season)

                    elif event == 'Start Flask Server':
                        commands = 'cd C:\\Users\\Roy\\Documents\\GitHub\\vfhl\\python && C:\\Users\\Roy\\AppData\\Local\\Programs\\Python\\Python310\\python.exe main.py'
                        # Execute the commands
                        subprocess.Popen('cmd.exe /K ' + commands)

                    elif event == 'Start Daily VFHL Scheduled Task':
                        self.start_daily_vfhl_scheduled_task()

                    elif event == 'Get Pool Standings':
                        self.get_pool_standings()
                        update_pool_teams_tab = True
                        refresh_pool_team_config = True

                except Exception as e:
                    sg.popup_error(f'Error in {sys._getframe().f_code.co_name}: {e}')
                    # set these to False, to prevent eternal looping
                    refresh_pool_team_config = False
                    refresh_pool_team_roster_config = False
                    continue

        except Exception as e:
            sg.PopupScrolled(''.join(traceback.format_exception(type(e), value=e, tb=e.__traceback__)), modal=True)
            pass

        finally:
            try:
                window.Close()
            except Exception as e:
                sg.PopupScrolled(''.join(traceback.format_exception(type(e), value=e, tb=e.__traceback__)), modal=True)
                pass

        return

    def writeToExcel(self, excelFile: str='', excelSheets: list=[], batch: bool=False, logger: logging.Logger=None, dialog: sg.Window=None):
        """
        Write multiple DataFrames to an Excel file based on a list of excelSheets.

        Args:
            excelSheets (list): A list of dictionaries, each containing the arguments for a write operation.
            excelFile (str): The name of the Excel file.
            batch (bool): Whether the operation is part of a batch process.
            logger (logging.Logger): Logger for batch operations.
            dialog (sg.Window): Dialog for progress updates.
        """
        try:
            if excelFile == '' or len(excelSheets) == 0:
                msg = 'No excel file or sheets to write to...'
                if batch:
                    logger.debug(msg)
                else:
                    dialog['-PROG-'].update(msg)
                    event, values = dialog.read(timeout=2)
                    if event == 'Cancel' or event == sg.WIN_CLOSED:
                        return

            msg = f'Writing to Excel...'
            if batch:
                logger.debug(msg)
            else:
                dialog['-PROG-'].update(msg)
                event, values = dialog.read(timeout=2)
                if event == 'Cancel' or event == sg.WIN_CLOSED:
                    return

            # Path to the OneDrive file
            one_drive_file = f"C:/Users/Roy/OneDrive/VFHL/{excelFile}"

            # Load the existing workbook
            workbook = openpyxl.load_workbook(one_drive_file)

            try:
                for task in excelSheets:
                    df = task.get('df')
                    excelSheet = task.get('excelSheet')
                    min_row = task.get('min_row', 2)
                    max_row = task.get('max_row', None)
                    min_col = task.get('min_col', 1)

                    # Select the target sheet
                    sheet = workbook[excelSheet]

                    max_df_col = len(df.columns)

                    if max_row is None:
                        max_row = sheet.max_row

                    # Clear existing data in the target range
                    for row in sheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_df_col):
                        for cell in row:
                            cell.value = None

                    # Write the DataFrame to the sheet starting at the specified row and column
                    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), min_row):
                        for c_idx, value in enumerate(row[:max_df_col], min_col):
                            sheet.cell(row=r_idx, column=c_idx, value=value)

                # Save the workbook
                workbook.save(one_drive_file)

            finally:
                # Ensure the workbook is closed
                workbook.close()

            # Trigger OneDrive sync for the xlsx file
            one_drive_file = one_drive_file.replace('/', '\\').replace("'", "''")
            result = subprocess.run(
                ["powershell", "-Command", f'Start-Sleep -Seconds 5; Start-Process -FilePath "C:\\Users\\Roy\\AppData\\Local\\Microsoft\\OneDrive\\OneDrive.exe" -ArgumentList "/sync {one_drive_file} /background"'],
                capture_output=True,
                text=True
            )

            if result.returncode != 0:
                error_msg = f"Error syncing OneDrive file: {result.stderr}"
                if batch:
                    logger.error(error_msg)
                else:
                    sg.popup_error(error_msg)

        except Exception as e:
            msg = f'Error in {sys._getframe().f_code.co_name}: {e}'
            if batch:
                logger.error(msg)
            else:
                sg.popup_error(msg)

        finally:
            msg = 'Write to Excel completed...'
            if batch:
                logger.debug(msg)
            else:
                sg.popup_notify(msg, title=sys._getframe().f_code.co_name)

        return

class ProgressBar():

    def __init__(self, max_value: int):

        self.layout = [
            [sg.Text('', key='progbar_text')],
            [sg.ProgressBar(max_value, orientation='h', size=(20, 20), key='progbar')],
            [sg.Cancel(pad=(5,5))],
        ]

        return

def main():

    try:

        hp = HockeyPool()

        global season
        seasons = Season().getCurrent()
        if len(seasons) == 0:
            return

        season = seasons[0] # if there multiple, it doesn't matter at this point since only interested in season id
        season.set_season_constants()

        with get_db_connection() as connection:
            cursor = connection.cursor()

            sql = f'select * from HockeyPool hp where season_id={season.id}'
            cursor.execute(sql)
            rows = cursor.fetchall()
            if rows:
                row = rows[0]

                for key in row.keys():
                    setattr(hp, key, row[key])

        cursor.close()
        connection.close()

        hp.window()

    except Exception as e:
        sg.PopupScrolled(''.join(traceback.format_exception(type(e), value=e, tb=e.__traceback__)), modal=True)

    return

if __name__ == "__main__":

    main()

    exit()
